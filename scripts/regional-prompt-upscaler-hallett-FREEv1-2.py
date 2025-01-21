# Regional Prompt Upscaler - Free Version v1.12
#
# This free version removes the following features:
# 1) Exclude Humans/Animals
# 2) Turn to Night
# 3) Exclude Vehicles (Except Boats)
# 4) Enable Blur Detection
# 5) Words to Remove
# 6) Categories to Remove
#
# The corresponding UI elements remain in a disabled/greyed-out state labeled "Pro Only".
# "Included with purchase of LoRa model from hallett-ai.com"

import subprocess
import sys
import hashlib
import io

# Utility functions to install missing packages
def install(package):
    try:
        print(f"Installing package: {package}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    except Exception as e:
        print(f"Failed to install {package}: {e}")

def ensure_package_installed(package, module_name=None):
    if module_name is None:
        module_name = package
    try:
        __import__(module_name)
    except ImportError:
        print(f"{package} is not installed. Attempting to install...")
        install(package)

# Ensure required packages are installed before imports
def install_dependencies():
    ensure_package_installed('torch')
    ensure_package_installed('transformers')
    ensure_package_installed('openpyxl')
    ensure_package_installed('numpy')
    ensure_package_installed('opencv-python', module_name='cv2')
    ensure_package_installed('Pillow', module_name='PIL')
    ensure_package_installed('gradio')
    ensure_package_installed('clip-interrogator')

install_dependencies()

# Now imports
import torch
import cv2  # OpenCV, used here only if needed for expansions (kept in case you restore features)
import math
import gradio as gr
import numpy as np
from PIL import Image, ImageDraw, ImageFilter
from modules import processing, shared, images, devices, scripts
from modules.processing import Processed, fix_seed
from modules.shared import opts, state
import os
import re
import copy
import gc

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from transformers import (
    BlipProcessor, BlipForConditionalGeneration,
    ViTImageProcessor, GPT2Tokenizer, VisionEncoderDecoderModel,
    AutoProcessor, AutoModelForCausalLM
)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

###############################################################################
# Utility function to format LoRA in prefix
###############################################################################
def set_lora_in_prefix(lora_selection):
    """
    Formats the selected LoRA model for insertion into the Prefix box.
    Removes the file extension and adds <lora:filename:1>, 
    including a trailing comma + space.
    """
    if lora_selection is None or lora_selection == "None":
        return ""
    filename = os.path.splitext(os.path.basename(lora_selection))[0]
    return f"<lora:{filename}:1>, "

###############################################################################
# Prompt Generators
###############################################################################
class PromptGeneratorBase:
    def __init__(self, device: torch.device):
        self.device = device

    def generate_prompt(self, image: Image.Image, prompt_mode: str = "Complex",
                        max_words: int = None, prefix: str = "", suffix: str = "") -> str:
        raise NotImplementedError("generate_prompt method must be implemented by subclasses.")

    def supports_save_load(self) -> bool:
        return True


class BLIPPromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device, model_name: str = "Salesforce/blip-image-captioning-base"):
        super().__init__(device)
        self.processor = BlipProcessor.from_pretrained(model_name)
        self.model = BlipForConditionalGeneration.from_pretrained(
            model_name,
            torch_dtype=torch.float16 if device.type == 'cuda' else torch.float32
        ).to(self.device)
        self.model.eval()

    def generate_prompt(self, image: Image.Image, prompt_mode: str = "Complex",
                        max_words: int = None, prefix: str = "", suffix: str = "") -> str:
        try:
            if prompt_mode.lower() == "complex":
                prompt_text = "Describe this image in detail."
            else:
                prompt_text = "A picture of"

            inputs = self.processor(
                images=image,
                text=prompt_text,
                return_tensors="pt"
            ).to(self.device, torch.float16 if self.device.type == 'cuda' else torch.float32)

            with torch.inference_mode():
                if self.device.type == "cuda":
                    with torch.cuda.amp.autocast():
                        outputs = self.model.generate(
                            **inputs,
                            max_length=100 if prompt_mode.lower() == "complex" else 50,
                            num_beams=5 if prompt_mode.lower() == "complex" else 1,
                            no_repeat_ngram_size=2,
                            early_stopping=True
                        )
                else:
                    outputs = self.model.generate(
                        **inputs,
                        max_length=100 if prompt_mode.lower() == "complex" else 50,
                        num_beams=5 if prompt_mode.lower() == "complex" else 1,
                        no_repeat_ngram_size=2,
                        early_stopping=True
                    )

            caption = self.processor.decode(outputs[0], skip_special_tokens=True)

            if max_words:
                caption = self._trim_caption(caption, max_words=max_words)
            else:
                if prompt_mode.lower() == "simple":
                    caption = self._trim_caption(caption, min_words=5, max_words=15)

            prompt = f"{prefix.strip()} {caption} {suffix.strip()}".strip()
            return prompt

        except Exception as e:
            print(f"Error generating prompt with BLIP: {e}")
            return "an image"

        finally:
            del inputs, outputs
            torch.cuda.empty_cache()
            gc.collect()

    def _trim_caption(self, caption: str, min_words: int = 0, max_words: int = None) -> str:
        words = caption.split()
        if max_words and len(words) > max_words:
            caption = " ".join(words[:max_words])
        if min_words and len(words) < min_words:
            pass
        return caption


class ViTGPT2PromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device):
        super().__init__(device)
        self.image_processor = ViTImageProcessor.from_pretrained("nlpconnect/vit-gpt2-image-captioning")
        self.tokenizer = GPT2Tokenizer.from_pretrained("nlpconnect/vit-gpt2-image-captioning")
        self.model = VisionEncoderDecoderModel.from_pretrained(
            "nlpconnect/vit-gpt2-image-captioning"
        ).to(self.device).eval()

    def generate_prompt(self, image: Image.Image, prompt_mode: str = "Complex",
                        max_words: int = None, prefix: str = "", suffix: str = "") -> str:
        pixel_values = self.image_processor(images=image, return_tensors="pt").pixel_values.to(self.device)

        num_beams = 5 if prompt_mode.lower() == "complex" else 1
        max_length = 50 if prompt_mode.lower() == "complex" else 20
        repetition_penalty = 1.2 if prompt_mode.lower() == "complex" else 1.0

        with torch.no_grad():
            output_ids = self.model.generate(
                pixel_values,
                max_length=max_length,
                num_beams=num_beams,
                repetition_penalty=repetition_penalty,
                no_repeat_ngram_size=2 if prompt_mode.lower() == "complex" else 0,
                early_stopping=True
            )
        caption = self.tokenizer.decode(output_ids[0], skip_special_tokens=True)

        if max_words:
            caption = " ".join(caption.split()[:max_words])

        prompt = f"{prefix.strip()} {caption} {suffix.strip()}".strip()
        return prompt


class Florence2PromptGenerator(PromptGeneratorBase):
    """
    Wraps Microsoft Florence-2, skipping flash_attn references.
    """
    def __init__(self, device: torch.device, task_option: str = "Caption"):
        super().__init__(device)
        self.device = device
        self.model_id = 'microsoft/Florence-2-large'

        from transformers.dynamic_module_utils import get_imports
        from unittest.mock import patch

        def fixed_get_imports(filename):
            # Avoid flash_attn trouble
            if not str(filename).endswith("modeling_florence2.py"):
                return get_imports(filename)
            imports = get_imports(filename)
            if "flash_attn" in imports:
                imports.remove("flash_attn")
            return imports

        with patch("transformers.dynamic_module_utils.get_imports", fixed_get_imports):
            self.processor = AutoProcessor.from_pretrained(self.model_id, trust_remote_code=True)
            self.model = AutoModelForCausalLM.from_pretrained(self.model_id, trust_remote_code=True).to(self.device).eval()

        torch.cuda.empty_cache()

    def generate_prompt(self, image: Image.Image, prompt_mode: str = "Complex",
                        max_words: int = None, prefix: str = "", suffix: str = "") -> str:
        task_prompt = self.get_task_prompt(prompt_mode)
        inputs = self.processor(text=task_prompt, images=image, return_tensors="pt").to(self.device)

        max_new_tokens = 50 if prompt_mode.lower() == "simple" else 1024

        with torch.no_grad():
            generated_ids = self.model.generate(
                input_ids=inputs["input_ids"],
                pixel_values=inputs["pixel_values"],
                max_new_tokens=max_new_tokens,
                early_stopping=True,
                do_sample=False,
                num_beams=3,
            )
        generated_text = self.processor.batch_decode(generated_ids, skip_special_tokens=False)[0]
        caption = self.process_caption(generated_text, task_prompt, image.size)

        if max_words:
            caption = " ".join(caption.split()[:max_words])

        prompt = f"{prefix.strip()} {caption} {suffix.strip()}".strip()
        return prompt

    def get_task_prompt(self, prompt_mode: str) -> str:
        if prompt_mode.lower() == "simple":
            return "<CAPTION>"
        else:
            return "<DETAILED_CAPTION>"

    def process_caption(self, generated_text, task_prompt, image_size):
        parsed_answer = self.processor.post_process_generation(
            generated_text,
            image_size=image_size,
            task=task_prompt
        )
        caption = parsed_answer.get(task_prompt, "")
        if caption.lower().startswith("this image is"):
            caption = caption[len("This image is"):].strip()
        return caption


class CLIPPromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device):
        super().__init__(device)
        from clip_interrogator import Config, Interrogator
        self.device = device
        self.config = Config(
            clip_model_name="ViT-L-14/openai",
            device="cuda" if device.type == "cuda" else "cpu"
        )
        self.interrogator = Interrogator(self.config)
        self.prompt_cache = {}

    def generate_prompt(self, image: Image.Image, prompt_mode: str = "Complex",
                        max_words: int = None, prefix: str = "", suffix: str = "") -> str:
        try:
            image_hash = self.generate_image_hash(image)
            if image_hash in self.prompt_cache:
                return self.prompt_cache[image_hash]

            if prompt_mode.lower() == "simple":
                caption = self.interrogator.interrogate_fast(image)
                caption = self._trim_caption(caption, min_words=5, max_words=15)
            else:
                caption = self.interrogator.interrogate(image)
                caption = self._trim_caption(caption, max_words=50)

            if max_words:
                caption = self._trim_caption(caption, max_words=max_words)

            prompt = f"{prefix.strip()} {caption} {suffix.strip()}".strip()
            self.prompt_cache[image_hash] = prompt
            return prompt

        except Exception as e:
            print(f"Error generating prompt with CLIPInterrogator: {e}")
            return "an image"

        finally:
            torch.cuda.empty_cache()

    def generate_image_hash(self, image: Image.Image) -> str:
        with io.BytesIO() as buffer:
            image.save(buffer, format="PNG")
            image_bytes = buffer.getvalue()
        return hashlib.md5(image_bytes).hexdigest()

    def _trim_caption(self, caption: str, min_words: int = 0, max_words: int = None) -> str:
        words = caption.split()
        if max_words and len(words) > max_words:
            caption = " ".join(words[:max_words])
        if min_words and len(words) < min_words:
            pass
        return caption

###############################################################################
# USDURedraw (Free Version)
###############################################################################
class USDURedraw:
    """
    The heart of the tile-based upscaling approach.
    In the Free Version, all "Pro" features have been removed from logic
    (though disabled placeholders remain in UI).
    """
    def __init__(
        self,
        usdupscaler,
        device,
        feather_amount=50,
        save_tiles=False,
        prompt_generator=None
    ):
        self.upscaler = usdupscaler
        self.device = device
        self.save = False
        self.save_tiles = save_tiles
        self.tile_size = 1024
        self.enabled = True
        self.base_prompt = ""
        self.initial_info = None
        self.overlap_percentage = 0
        self.clip_prompt_suffix = ""
        self.entire_image_prompt = ""
        self.feather_amount = feather_amount
        self.prompt_generator = prompt_generator

        # Placeholders for Pro-Only features, no effect in free version:
        # (Kept just to show they exist, but they do nothing here.)
        # This script has stripped out the usage of these attributes:
        self.exclude_humans_animals = False
        self.words_to_remove = []
        self.categories_to_remove = []
        self.auto_adjust_tile_size = False
        self.turn_to_night = False
        self.exclude_vehicles = False
        self.blur_detection_enabled = False

    def calc_rectangle(self, xi, yi):
        x1 = self.tile_positions_x[xi]
        y1 = self.tile_positions_y[yi]
        if xi == self.cols - 1:
            x2 = self.width
        else:
            x2 = x1 + self.tile_size
        if yi == self.rows - 1:
            y2 = self.height
        else:
            y2 = y1 + self.tile_size

        print(f"Tile ({xi}, {yi}) position: ({x1}, {y1}, {x2}, {y2})")
        return x1, y1, x2, y2

    def analyze_full_image(self):
        """
        Generate a global prompt for the entire image if a prompt generator is set.
        """
        if self.prompt_generator is None:
            self.entire_image_prompt = self.upscaler.p.prompt or ""
            return
        print("Generating prompt for the entire image (Complex)...")
        self.entire_image_prompt = self.prompt_generator.generate_prompt(
            self.upscaler.image,
            prompt_mode="Complex",
            prefix=self.base_prompt,
            suffix=self.clip_prompt_suffix
        )
        print(f"Entire image prompt: {self.entire_image_prompt}")

    def generate_tile_prompt(self, tile_image):
        """
        Generate the tile prompt using the selected prompt generator.
        """
        if self.prompt_generator is None:
            tile_prompt = f"{self.upscaler.p.prompt}"
        else:
            print("Generating prompt for tile...")
            tile_prompt = self.prompt_generator.generate_prompt(
                tile_image,
                prompt_mode=self.upscaler.prompt_mode_tile,
                prefix=self.base_prompt,
                suffix=self.clip_prompt_suffix
            )
        return tile_prompt

    def pre_analyze(self, image):
        """
        Pre-analyzes the entire image and each tile to produce prompts.
        """
        self.analyze_full_image()
        tile_prompts = []
        for yi in range(self.rows):
            row_prompts = []
            for xi in range(self.cols):
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)
                x1 = max(0, min(x1, self.width))
                y1 = max(0, min(y1, self.height))
                x2 = max(0, min(x2, self.width))
                y2 = max(0, min(y2, self.height))
                if x2 <= x1 or y2 <= y1:
                    continue

                tile_image = image.crop((x1, y1, x2, y2))
                tile_prompt = self.generate_tile_prompt(tile_image)
                row_prompts.append(tile_prompt)
            tile_prompts.append(row_prompts)
        return tile_prompts

    def start(self, p, image, tile_prompts):
        """
        Processes each tile with the relevant prompt, reassembles the final image.
        """
        self.width, self.height = image.size
        final_image = Image.new("RGB", (self.width, self.height))

        for yi in range(self.rows):
            for xi in range(self.cols):
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)
                x1 = max(0, min(x1, self.width))
                y1 = max(0, min(y1, self.height))
                x2 = max(0, min(x2, self.width))
                y2 = max(0, min(y2, self.height))
                if x2 <= x1 or y2 <= y1:
                    continue

                tile_image = image.crop((x1, y1, x2, y2))
                tile_prompt = tile_prompts[yi][xi]

                print(f"Tile ({xi}, {yi}) prompt: {tile_prompt}")
                tile_p = self.prepare_tile_processing(p, tile_image, tile_prompt)
                processed = processing.process_images(tile_p)
                processed_tile = processed.images[0]
                processed_tile = processed_tile.resize(tile_image.size, Image.LANCZOS)

                if self.save_tiles:
                    tile_filename = f"tile_{yi}_{xi}.png"
                    tile_filepath = os.path.join(p.outpath_samples, tile_filename)
                    processed_tile.save(tile_filepath)
                    print(f"Saved tile: {tile_filepath}")

                # Feather if user set > 0
                if self.feather_amount > 0:
                    mask = self.create_feathered_mask((x1, y1, x2, y2), (self.width, self.height), self.feather_amount)
                else:
                    mask = Image.new("L", tile_image.size, 255)

                if mask.size != tile_image.size:
                    mask = mask.resize(tile_image.size, Image.LANCZOS)

                final_image.paste(processed_tile, (x1, y1), mask)

                if self.device.type == 'cuda':
                    torch.cuda.empty_cache()
                del tile_image, processed, processed_tile, mask, tile_p
                gc.collect()

                state.job_no += 1
                state.job_count = self.upscaler.job_count
                if state.interrupted:
                    break
            if state.interrupted:
                break

        return final_image

    def create_feathered_mask(self, tile_position, image_size, feather_radius):
        """
        Creates a feathered mask so that each tile transitions smoothly
        into its neighboring area.
        """
        tile_width = tile_position[2] - tile_position[0]
        tile_height = tile_position[3] - tile_position[1]
        x1, y1, x2, y2 = tile_position
        image_width, image_height = image_size

        mask = Image.new("L", (tile_width, tile_height), 255)
        needs_feathering = False
        left = 0
        top = 0
        right = tile_width
        bottom = tile_height

        if x1 > 0:
            left = feather_radius
            needs_feathering = True
        if y1 > 0:
            top = feather_radius
            needs_feathering = True
        if x2 < image_width:
            right = tile_width - feather_radius
            needs_feathering = True
        if y2 < image_height:
            bottom = tile_height - feather_radius
            needs_feathering = True

        if needs_feathering:
            alpha = Image.new("L", (tile_width, tile_height), 0)
            draw = ImageDraw.Draw(alpha)
            draw.rectangle((left, top, right, bottom), fill=255)
            mask = alpha.filter(ImageFilter.GaussianBlur(radius=feather_radius))
        else:
            mask = Image.new("L", (tile_width, tile_height), 255)

        return mask

    def prepare_tile_processing(self, p, tile_image, tile_prompt):
        """
        Sets up the processing for a single tile (width/height prompt, etc).
        """
        tile_p = copy.copy(p)
        tile_p.width = min(tile_image.width, max(64, (tile_image.width // 64) * 64))
        tile_p.height = min(tile_image.height, max(64, (tile_image.height // 64) * 64))

        resized_tile_image = tile_image.resize((tile_p.width, tile_p.height), Image.LANCZOS)
        tile_p.init_images = [resized_tile_image]
        tile_p.prompt = tile_prompt
        tile_p.batch_size = 1
        tile_p.n_iter = 1
        tile_p.do_not_save_samples = not self.save_tiles
        tile_p.do_not_save_grid = True

        fix_seed(tile_p)
        return tile_p

###############################################################################
# USDUpscaler (Free Version)
###############################################################################
class USDUpscaler:
    model_cache = {}

    def __init__(
        self,
        p,
        image,
        upscaler_index: int,
        save_redraw,
        tile_size,
        base_prompt,
        overlap_percentage,
        clip_prompt_suffix,
        prompt_method: str,
        prompt_mode_tile: str,
        save_prompts_to_file,
        uploaded_prompt_file,
        global_seed: int,
        feather_amount: int,
        save_tiles: bool
    ):
        self.p = p
        self.image: Image = image
        self.prompt_mode_tile = prompt_mode_tile

        # Calculate scale factor from the difference between original vs. p.width/p.height
        init_img_width, init_img_height = self.image.size
        desired_width, desired_height = p.width, p.height
        scale_factor_w = desired_width / init_img_width
        scale_factor_h = desired_height / init_img_height
        if abs(scale_factor_w - scale_factor_h) > 0.01:
            print("Warning: Aspect ratio has changed. Using width scale factor.")
        self.scale_factor = scale_factor_w

        self.upscaler_data = shared.sd_upscalers[upscaler_index]
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

        # Initialize the chosen prompt generator
        self.prompt_method = prompt_method
        self.prompt_generator = self.initialize_prompt_generator(prompt_method)

        # Setup the tile-based redraw object
        self.redraw = USDURedraw(
            usdupscaler=self,
            device=self.device,
            feather_amount=feather_amount,
            save_tiles=save_tiles,
            prompt_generator=self.prompt_generator
        )

        self.redraw.save = save_redraw
        self.redraw.base_prompt = base_prompt
        self.redraw.clip_prompt_suffix = clip_prompt_suffix
        self.initial_info = None

        self.save_prompts_to_file = save_prompts_to_file
        self.uploaded_prompt_file = uploaded_prompt_file
        self.global_seed = global_seed

        # Perform the upscaling
        self.upscale_image()

        # Adjust tile positions
        self.adjust_tile_size_and_positions(tile_size, overlap_percentage)

    def initialize_prompt_generator(self, method_name: str):
        if method_name == "NONE":
            return None
        if method_name in USDUpscaler.model_cache:
            return USDUpscaler.model_cache[method_name]

        try:
            if method_name == "BLIP":
                generator = BLIPPromptGenerator(self.device)
            elif method_name == "ViT-GPT2":
                generator = ViTGPT2PromptGenerator(self.device)
            elif method_name == "Florence-2":
                generator = Florence2PromptGenerator(self.device, task_option="Detailed Caption")
            elif method_name == "CLIP":
                generator = CLIPPromptGenerator(self.device)
            else:
                raise ValueError(f"Unsupported prompt method: {method_name}")

            USDUpscaler.model_cache[method_name] = generator
            return generator
        except Exception as e:
            print(f"Error initializing prompt generator '{method_name}': {e}")
            raise e

    def upscale_image(self):
        if self.upscaler_data:
            print(f"Upscaling image using {self.upscaler_data.name}...")
            self.image = self.upscaler_data.scaler.upscale(
                self.image, 
                self.scale_factor, 
                self.upscaler_data.data_path
            )
            self.p.width, self.p.height = self.image.size
            print(f"Image upscaled to {self.p.width}x{self.p.height}")
        else:
            print("No upscaler selected. Skipping upscaling step.")

    def adjust_tile_size_and_positions(self, tile_size, overlap_percentage):
        """
        Calculates how many columns/rows we need, sets up the coordinate arrays.
        (Free version uses tile_size as is, no auto-size adjusting.)
        """
        width, height = self.image.size
        self.redraw.tile_size = tile_size
        print(f"Using specified tile size {self.redraw.tile_size}")

        self.overlap = int(overlap_percentage / 100 * self.redraw.tile_size)
        self.redraw.overlap_percentage = overlap_percentage

        self.cols = max(1, math.ceil((width + self.overlap) / (self.redraw.tile_size - self.overlap)))
        self.rows = max(1, math.ceil((height + self.overlap) / (self.redraw.tile_size - self.overlap)))

        # Build tile_positions_x
        self.redraw.tile_positions_x = []
        for i in range(self.cols):
            if i == self.cols - 1:
                pos = width - self.redraw.tile_size
            else:
                pos = i * (self.redraw.tile_size - self.overlap)
            self.redraw.tile_positions_x.append(pos)

        # Build tile_positions_y
        self.redraw.tile_positions_y = []
        for i in range(self.rows):
            if i == self.rows - 1:
                pos = height - self.redraw.tile_size
            else:
                pos = i * (self.redraw.tile_size - self.overlap)
            self.redraw.tile_positions_y.append(pos)

        self.redraw.cols = self.cols
        self.redraw.rows = self.rows
        self.redraw.width = width
        self.redraw.height = height

        print(f"Number of tiles: {self.cols} cols x {self.rows} rows")

    def pre_analyze_tiles(self):
        if self.uploaded_prompt_file:
            self.tile_prompts = self.load_prompts_from_excel(self.uploaded_prompt_file)
        else:
            self.tile_prompts = self.redraw.pre_analyze(self.image)
            if self.save_prompts_to_file and self.prompt_generator is not None:
                self.save_prompts_to_excel(self.tile_prompts)

    def calc_jobs_count(self):
        self.job_count = self.redraw.rows * self.redraw.cols
        state.job_count = self.job_count

    def add_extra_info(self):
        self.initial_info = "Upscaling complete (Free Version)."

    def process(self):
        state.begin()
        self.calc_jobs_count()
        self.result_images = []

        self.pre_analyze_tiles()
        self.image = self.redraw.start(self.p, self.image, self.tile_prompts)
        self.initial_info = self.redraw.initial_info

        self.result_images.append(self.image)
        if self.redraw.save:
            print("Saving the final upscaled image...")
            self.save_image()

        self.finalize_processing()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

        state.end()

    def finalize_processing(self):
        if self.prompt_generator:
            del self.prompt_generator
        torch.cuda.empty_cache()
        gc.collect()

    def save_image(self):
        method_suffix = f"_{self.prompt_method}"
        images.save_image(
            self.result_images[0],
            self.p.outpath_samples,
            "",
            seed=self.global_seed,
            prompt=self.p.prompt,
            extension=opts.samples_format,
            info=self.initial_info,
            p=self.p,
            suffix=method_suffix
        )

    def save_prompts_to_excel(self, tile_prompts):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tile Prompts"

        for yi, row_prompts in enumerate(tile_prompts):
            for xi, prompt in enumerate(row_prompts):
                cell = ws.cell(row=yi+1, column=xi+1)
                cell.value = prompt
                cell.alignment = Alignment(wrap_text=True)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value) if column_cells else 0
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        excel_file_base = os.path.join(self.p.outpath_samples, f"tile_prompts_{self.prompt_method}")
        excel_file_path = excel_file_base + ".xlsx"
        suffix = 1
        while True:
            try:
                wb.save(excel_file_path)
                print(f"Tile prompts saved to {excel_file_path}")
                break
            except PermissionError:
                excel_file_path = excel_file_base + f"_{suffix}.xlsx"
                suffix += 1

    def load_prompts_from_excel(self, uploaded_file):
        if uploaded_file is None:
            raise ValueError("No Excel file uploaded.")
        wb = load_workbook(filename=uploaded_file.name)
        ws = wb.active

        tile_prompts = []
        for row in ws.iter_rows(values_only=True):
            row_prompts = [cell if cell is not None else "" for cell in row]
            tile_prompts.append(row_prompts)

        expected_rows = self.redraw.rows
        expected_cols = self.redraw.cols
        actual_rows = len(tile_prompts)
        actual_cols = max(len(r) for r in tile_prompts) if tile_prompts else 0
        if actual_rows != expected_rows or actual_cols != expected_cols:
            raise ValueError(
                f"Excel file dimensions ({actual_rows}x{actual_cols}) do not match "
                f"expected tile grid ({expected_rows}x{expected_cols})."
            )

        return tile_prompts

    # Recursively list subfolders for LoRA
    def get_lora_files(self):
        lora_path = os.path.join("models", "Lora")
        if not os.path.exists(lora_path):
            return ["None"]

        all_loras = []
        for root, dirs, files in os.walk(lora_path):
            for f in files:
                if f.lower().endswith(".safetensors") or f.lower().endswith(".ckpt"):
                    rel_path = os.path.relpath(os.path.join(root, f), lora_path)
                    rel_path = rel_path.replace(os.path.sep, "/")
                    entry = f"/LoRA/{rel_path}"
                    all_loras.append(entry)
        all_loras.sort()
        return ["None"] + all_loras


###############################################################################
# Script Class for Gradio (Free Version)
###############################################################################
class Script(scripts.Script):
    def title(self):
        return "Regional Prompt Upscaler - Free Version v1.12"

    def show(self, is_img2img):
        return is_img2img

    # Recursively list LoRA files (subfolders included)
    def get_lora_files(self):
        lora_path = os.path.join("models", "Lora")
        if not os.path.exists(lora_path):
            return ["None"]

        all_loras = []
        for root, dirs, files in os.walk(lora_path):
            for f in files:
                if f.lower().endswith(".safetensors") or f.lower().endswith(".ckpt"):
                    rel_path = os.path.relpath(os.path.join(root, f), lora_path)
                    rel_path = rel_path.replace(os.path.sep, "/")
                    entry = f"/LoRA/{rel_path}"
                    all_loras.append(entry)
        all_loras.sort()
        return ["None"] + all_loras

    def ui(self, is_img2img):
        """
        Builds the Gradio UI for the free version. 
        The "Pro Only" features are shown greyed out with a note.
        """

        elem_id_prefix = "regional_upscaler_free_v112"

        version_info = gr.HTML(
            "<p style=\"font-weight:bold;\">Regional Prompt Upscaler - Free Version v1.12</p>"
        )
        website_info = gr.HTML("<p><a href='https://hallett-ai.com' target='_blank'>hallett-ai.com</a></p>")
        free_message = gr.HTML("<p>This is the free version; certain advanced features are Pro Only.</p>")

        info = gr.HTML(
            "<p>Upscale your image in tiles. LoRA selection is auto-injected into the Prefix prompt if desired.</p>"
        )

        with gr.Row():
            upscaler_name = gr.Radio(
                label='Upscaler Model',
                elem_id=f"{elem_id_prefix}_upscaler_name",
                choices=[x.name for x in shared.sd_upscalers],
                value=shared.sd_upscalers[0].name,
            )

        with gr.Row():
            tile_size = gr.Slider(
                elem_id=f"{elem_id_prefix}_tile_size",
                minimum=512,
                maximum=2048,
                step=64,
                label='Tile size',
                value=1024
            )
            feather_amount = gr.Slider(
                elem_id=f"{elem_id_prefix}_feather_amount",
                label='Feather Amount',
                minimum=0,
                maximum=100,
                step=1,
                value=50
            )
            overlap_percentage = gr.Slider(
                label='Overlap Percentage',
                minimum=0,
                maximum=50,
                step=1,
                value=20
            )

        # Moved up in the UI: Prefix, Suffix
        gr.HTML("<p>Prompt Prefix and Suffix:</p>")
        with gr.Row():
            base_prompt = gr.Textbox(
                label="Prefix (Base Prompt)",
                placeholder="Enter a base prompt...",
                lines=1,
                value=""
            )
            clip_prompt_suffix = gr.Textbox(
                label="Suffix",
                placeholder="Enter a suffix for prompts...",
                lines=1
            )

        # Now LoRA selection (directly below Prefix/Suffix):
        gr.HTML("<p>LoRA Settings (Select LoRA → updates Prefix):</p>")
        with gr.Row():
            lora_model = gr.Dropdown(
                label="LoRA Model",
                choices=self.get_lora_files(),
                value="None"
            )
            refresh_lora_btn = gr.Button(value="Refresh LoRA List", variant="secondary")

        # On dropdown change, update 'base_prompt' with <lora:filename:1>,
        lora_model.change(
            lambda selection: gr.update(value=set_lora_in_prefix(selection)),
            inputs=[lora_model],
            outputs=[base_prompt]
        )

        refresh_lora_btn.click(
            lambda: gr.update(choices=self.get_lora_files()),
            inputs=[],
            outputs=[lora_model]
        )

        # Next: Prompt Generation Method
        prompt_methods = ["BLIP", "ViT-GPT2", "Florence-2", "CLIP", "NONE"]
        prompt_method = gr.Dropdown(
            label="Prompt Generation Method",
            choices=prompt_methods,
            value="Florence-2"
        )

        prompt_mode_tile = gr.Dropdown(
            label="Prompt Complexity for Tiles",
            choices=["Simple", "Complex"],
            value="Simple"
        )

        # Save options
        gr.HTML("<p>Save options:</p>")
        with gr.Row():
            save_prompts_to_file = gr.Checkbox(
                label='Save Prompts to Excel File',
                value=True
            )
            upload_prompt_file = gr.File(
                label='Upload Prompt Excel File (Optional)',
                file_types=['.xlsx']
            )

        with gr.Row():
            save_upscaled_image = gr.Checkbox(
                label="Save Upscaled Image",
                elem_id=f"{elem_id_prefix}_save_upscaled_image",
                value=True
            )
            save_tiles = gr.Checkbox(
                label="Save Each Tile",
                value=False
            )

        # -----------------------------------
        # Greyed-out Pro Features below:
        # -----------------------------------

        with gr.Row():
            exclude_humans_animals = gr.Checkbox(
                label='Exclude Humans/Animals (Pro Only)',
                value=False,
                interactive=False
            )
            exclude_vehicles = gr.Checkbox(
                label='Exclude Vehicles (Pro Only)',
                value=False,
                interactive=False
            )

        with gr.Row():
            turn_to_night = gr.Checkbox(
                label='Turn to Night (Pro Only)',
                value=False,
                interactive=False
            )
            blur_detection_enabled = gr.Checkbox(
                label='Enable Blur Detection (Pro Only)',
                value=False,
                interactive=False
            )

        with gr.Row():
            words_to_remove = gr.Textbox(
                label="Words to Remove (Pro Only)",
                placeholder="Pro Only",
                lines=1,
                interactive=False
            )
            categories_to_remove = gr.Textbox(
                label="Categories to Remove (Pro Only)",
                placeholder="Pro Only",
                lines=1,
                interactive=False
            )

        gr.HTML("<p style='margin-top: 1em;'>Included with purchase of LoRa model from "
                "<a href='https://hallett-ai.com' target='_blank'>hallett-ai.com</a></p>")
        gr.HTML("</div>")
        # -----------------------------------

        return [
            upscaler_name,
            tile_size,
            feather_amount,
            overlap_percentage,
            base_prompt,
            clip_prompt_suffix,
            prompt_method,
            prompt_mode_tile,
            save_prompts_to_file,
            upload_prompt_file,
            save_upscaled_image,
            save_tiles,
            # The next 6 are the "Pro only" placeholders, no effect:
            turn_to_night,
            exclude_humans_animals,
            words_to_remove,
            categories_to_remove,
            exclude_vehicles,
            blur_detection_enabled,
            lora_model
        ]

    def run(self, p, *args):
        (
            upscaler_name,
            tile_size,
            feather_amount,
            overlap_percentage,
            base_prompt,
            clip_prompt_suffix,
            prompt_method,
            prompt_mode_tile,
            save_prompts_to_file,
            upload_prompt_file,
            save_upscaled_image,
            save_tiles,
            # Pro-only placeholders (no effect in code):
            turn_to_night,
            exclude_humans_animals,
            words_to_remove,
            categories_to_remove,
            exclude_vehicles,
            blur_detection_enabled,
            lora_model_name
        ) = args

        # Find correct upscaler index
        upscaler_index = next((i for i, x in enumerate(shared.sd_upscalers) if x.name == upscaler_name), 0)

        fix_seed(p)
        global_seed = p.seed

        init_img = p.init_images[0]
        if init_img is None:
            return Processed(p, [], p.seed, "Empty image")

        init_img = images.flatten(init_img, opts.img2img_background_color)

        # Create the free upscaler instance
        upscaler = USDUpscaler(
            p=p,
            image=init_img,
            upscaler_index=upscaler_index,
            save_redraw=save_upscaled_image,
            tile_size=tile_size,
            base_prompt=base_prompt,
            overlap_percentage=overlap_percentage,
            clip_prompt_suffix=clip_prompt_suffix,
            prompt_method=prompt_method,
            prompt_mode_tile=prompt_mode_tile,
            save_prompts_to_file=save_prompts_to_file,
            uploaded_prompt_file=upload_prompt_file,
            global_seed=global_seed,
            feather_amount=feather_amount,
            save_tiles=save_tiles
        )

        upscaler.add_extra_info()
        upscaler.process()

        result_images = upscaler.result_images
        return Processed(p, result_images, p.seed, upscaler.initial_info or "")
