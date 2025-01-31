"""
Regional Prompt Upscaler and Detailer v1.3

Changes from 1.2:
1) Restores "Remove Words" textbox for additional user-defined words.
2) Fixes categories to a specific order: Humans, Animals, Vehicles (except boats), Sky, Water, Tech, Food, Plants.
3) Renames the "technology" category to "tech" in the code to match the requested label.

Usage:
- Place `categories.yaml` (as shown above) in the same folder.
- This script auto-generates checkboxes only for these 8 categories in that order.
- The 'Remove Words' text box is a separate user-provided list, also removed from prompts.
"""

import logging
import os
import subprocess
import sys
import re
import copy
import gc
import math
import hashlib
import io

from logging.handlers import RotatingFileHandler
from typing import List, Dict, Optional, Tuple

import torch
import cv2
import gradio as gr
from PIL import Image, ImageDraw, ImageFilter

# A1111 imports
from modules import processing, shared, images, scripts
from modules.processing import Processed, fix_seed
from modules.shared import opts, state

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from transformers import (
    BlipProcessor, BlipForConditionalGeneration,
    ViTImageProcessor, GPT2Tokenizer, VisionEncoderDecoderModel,
    AutoProcessor, AutoModelForCausalLM
)

from scripts.dof_features import DepthOfFieldFeatures


# -------------------------------------------------
# Logging
# -------------------------------------------------
def setup_logging(debug_enabled: bool = False):
    """Configure logging with both file and console handlers."""
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(log_dir, exist_ok=True)

    log_file = os.path.join(log_dir, 'regional_upscaler.log')

    file_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    console_formatter = logging.Formatter('%(levelname)s: %(message)s')

    file_handler = RotatingFileHandler(log_file, maxBytes=10 * 1024 * 1024, backupCount=5)
    file_handler.setFormatter(file_formatter)
    file_handler.setLevel(logging.DEBUG)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(console_formatter)
    console_handler.setLevel(logging.DEBUG if debug_enabled else logging.INFO)

    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG if debug_enabled else logging.INFO)
    # Clear old handlers if re-running
    if logger.hasHandlers():
        logger.handlers.clear()
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger

logger = setup_logging(debug_enabled=False)


# -------------------------------------------------
# Dependency Checking
# -------------------------------------------------
def install(package: str) -> bool:
    logger.info(f"Installing package: {package}")
    try:
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", package],
            capture_output=True,
            text=True,
            check=True
        )
        logger.debug(f"Installation output: {result.stdout}")
        return True
    except subprocess.CalledProcessError as e:
        logger.error(f"Failed to install {package}. Error: {e}")
        logger.debug(f"Installation error output: {e.stderr}")
        return False
    except Exception as e:
        logger.error(f"Unexpected error installing {package}: {e}")
        return False

def ensure_package_installed(package: str, module_name: Optional[str] = None) -> bool:
    if module_name is None:
        module_name = package
    try:
        __import__(module_name)
        logger.debug(f"Package {package} is already installed.")
        return True
    except ImportError:
        logger.info(f"Package {package} not found. Attempting installation...")
        return install(package)

def install_dependencies() -> bool:
    required_packages = [
        ('torch', None),
        ('transformers', None),
        ('openpyxl', None),
        ('opencv-python', 'cv2'),
        ('Pillow', 'PIL'),
        ('gradio', None),
        ('clip-interrogator', None),
    ]

    logger.info("Checking and installing required dependencies...")
    failed_installs = []
    for package, module_name in required_packages:
        if not ensure_package_installed(package, module_name):
            failed_installs.append(package)

    if failed_installs:
        logger.critical(f"Failed to install required packages: {', '.join(failed_installs)}")
        return False

    logger.info("All dependencies installed successfully.")
    return True

# Attempt to install any missing dependencies
if not install_dependencies():
    logger.critical("Some dependencies could not be installed. The script may fail.")


# -------------------------------------------------
# Night Mode
# -------------------------------------------------
import importlib.util

def load_night_mode():
    """Load night_mode.py if available, returning (bool_available, function_convert)."""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        potential_paths = [
            os.path.join(script_dir, "Latest", "night_mode.py"),
            os.path.join(script_dir, "night_mode.py"),
            os.path.join(os.path.dirname(script_dir), "Latest", "night_mode.py"),
        ]

        night_mode_file = None
        for path in potential_paths:
            if os.path.exists(path):
                night_mode_file = path
                logger.info(f"Found night_mode.py at: {path}")
                break

        if not night_mode_file:
            logger.warning("night_mode.py not found; night mode disabled.")
            return False, lambda x: x

        module_dir = os.path.dirname(night_mode_file)
        if module_dir not in sys.path:
            sys.path.insert(0, module_dir)

        if 'night_mode' in sys.modules:
            del sys.modules['night_mode']

        spec = importlib.util.spec_from_file_location("night_mode", night_mode_file)
        if not spec:
            logger.error("Could not create module spec for night_mode.py")
            return False, lambda x: x

        night_mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(night_mod)

        if not hasattr(night_mod, 'convert_to_night_prompt'):
            logger.error("convert_to_night_prompt not found in night_mode.py")
            return False, lambda x: x

        test_result = night_mod.convert_to_night_prompt("A sunny day")
        if not isinstance(test_result, str) or test_result == "A sunny day":
            logger.warning("Night mode function tested but returned unexpected or identical result.")
            return False, lambda x: x

        logger.info("night_mode.py loaded successfully.")
        return True, night_mod.convert_to_night_prompt

    except Exception as e:
        logger.error(f"Failed to load night_mode.py: {str(e)}")
        return False, lambda x: x

NIGHT_MODE_AVAILABLE, convert_to_night_prompt = load_night_mode()


# -------------------------------------------------
# Categories YAML
# -------------------------------------------------
def read_categories_yaml() -> Dict[str, List[str]]:
    """
    Reads 'categories.yaml' from the same folder, returns a dictionary:
    { 'humans': [...], 'animals': [...], etc. }
    Each category is a list of words (lowercased).
    """
    import yaml

    script_dir = os.path.dirname(os.path.abspath(__file__))
    cfg_path = os.path.join(script_dir, "categories.yaml")

    if not os.path.exists(cfg_path):
        logger.warning("No categories.yaml found; using empty fallback.")
        return {}

    try:
        with open(cfg_path, "r", encoding="utf-8") as f:
            parsed = yaml.safe_load(f)
        if not parsed or "categories" not in parsed:
            logger.warning("categories.yaml missing 'categories' key; using empty fallback.")
            return {}

        raw_cats: dict = parsed.get("categories", {})
        cat_dict: Dict[str, List[str]] = {}

        for cat_name, cat_string in raw_cats.items():
            if not cat_string:
                cat_dict[cat_name.lower().strip()] = []
                continue
            terms = [w.strip().lower() for w in cat_string.split(",") if w.strip()]
            cat_dict[cat_name.lower().strip()] = terms

        return cat_dict

    except Exception as e:
        logger.error(f"Error reading categories.yaml: {e}")
        return {}


# -------------------------------------------------
# LoRA Helper
# -------------------------------------------------
def set_lora_in_prefix(lora_selection: str) -> str:
    """Return `<lora:filename:1>, ` if user chooses a LoRA file."""
    if not lora_selection or lora_selection == "None":
        return ""
    filename = os.path.splitext(os.path.basename(lora_selection))[0]
    return f"<lora:{filename}:1>, "


# -------------------------------------------------
# Prompt Generators
# -------------------------------------------------
MAX_PROMPT_LENGTH = 150
MAX_SIMPLE_PROMPT_LENGTH = 50

class PromptGeneratorBase:
    def __init__(self, device: torch.device):
        self.device = device
        self.prompt_cache = {}

    def get_generation_config(self, prompt_mode: str) -> dict:
        if prompt_mode == "Complex":
            return {
                "max_length": MAX_PROMPT_LENGTH,
                "num_beams": 7,
                "no_repeat_ngram_size": 3,
                "temperature": 1.2,
                "do_sample": True,
                "early_stopping": True,
                "length_penalty": 1.5,
                "top_k": 50,
                "top_p": 0.95
            }
        else:
            return {
                "max_length": MAX_SIMPLE_PROMPT_LENGTH,
                "num_beams": 1,
                "no_repeat_ngram_size": 2,
                "temperature": 0.7,
                "do_sample": False,
                "early_stopping": True,
                "length_penalty": 1.0,
                "top_k": 0,
                "top_p": 1.0
            }

    def generate_prompt(
        self,
        image: Image.Image,
        prompt_mode: str = "Complex",
        max_words: Optional[int] = None,
        prefix: str = "",
        suffix: str = ""
    ) -> str:
        return f"{prefix.strip()} an image {suffix.strip()}".strip()


class BLIPPromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device, model_name="Salesforce/blip-image-captioning-base"):
        super().__init__(device)
        self.processor = BlipProcessor.from_pretrained(model_name)
        self.model = BlipForConditionalGeneration.from_pretrained(
            model_name,
            torch_dtype=torch.float16 if device.type == 'cuda' else torch.float32
        ).to(device).eval()

    def generate_prompt(
        self,
        image: Image.Image,
        prompt_mode: str = "Complex",
        max_words: Optional[int] = None,
        prefix: str = "",
        suffix: str = ""
    ) -> str:
        if prompt_mode.lower() == "complex":
            initial_text = "Provide a highly detailed description of the image, describing style, content, and context."
        else:
            initial_text = "A picture of"

        inputs = self.processor(
            images=image,
            text=initial_text,
            return_tensors="pt"
        ).to(self.device, dtype=torch.float16 if self.device.type == 'cuda' else torch.float32)

        config = self.get_generation_config(prompt_mode)
        try:
            with torch.no_grad():
                out_ids = self.model.generate(**inputs, **config)
            caption = self.processor.decode(out_ids[0], skip_special_tokens=True)
            if max_words:
                caption = " ".join(caption.split()[:max_words])

            return f"{prefix.strip()} {caption} {suffix.strip()}".strip()

        except Exception as e:
            logger.error(f"[BLIP Error] {e}")
            return "an image"
        finally:
            torch.cuda.empty_cache()
            gc.collect()


class ViTGPT2PromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device):
        super().__init__(device)
        self.image_processor = ViTImageProcessor.from_pretrained("nlpconnect/vit-gpt2-image-captioning")
        self.tokenizer = GPT2Tokenizer.from_pretrained("nlpconnect/vit-gpt2-image-captioning")
        self.model = VisionEncoderDecoderModel.from_pretrained("nlpconnect/vit-gpt2-image-captioning").to(device).eval()

    def generate_prompt(
        self,
        image: Image.Image,
        prompt_mode: str = "Complex",
        max_words: Optional[int] = None,
        prefix: str = "",
        suffix: str = ""
    ) -> str:
        config = self.get_generation_config(prompt_mode)
        pixel_values = self.image_processor(images=image, return_tensors="pt").pixel_values.to(self.device)
        try:
            with torch.no_grad():
                out = self.model.generate(pixel_values, **config)
            caption = self.tokenizer.decode(out[0], skip_special_tokens=True)
            if max_words:
                caption = " ".join(caption.split()[:max_words])
            return f"{prefix.strip()} {caption} {suffix.strip()}".strip()

        except Exception as e:
            logger.error(f"[ViT-GPT2 Error] {e}")
            return "an image"


class Florence2PromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device, task_option="Caption"):
        super().__init__(device)
        self.model_id = "microsoft/Florence-2-large"
        from transformers.dynamic_module_utils import get_imports
        from unittest.mock import patch

        def fixed_get_imports(filename):
            if not str(filename).endswith("modeling_florence2.py"):
                return get_imports(filename)
            imports = get_imports(filename)
            if "flash_attn" in imports:
                imports.remove("flash_attn")
            return imports

        with patch("transformers.dynamic_module_utils.get_imports", fixed_get_imports):
            self.processor = AutoProcessor.from_pretrained(self.model_id, trust_remote_code=True)
            self.model = AutoModelForCausalLM.from_pretrained(self.model_id, trust_remote_code=True).to(device).eval()

    def generate_prompt(
        self,
        image: Image.Image,
        prompt_mode: str = "Complex",
        max_words: Optional[int] = None,
        prefix: str = "",
        suffix: str = ""
    ) -> str:
        if prompt_mode.lower() == "simple":
            task_prompt = "<CAPTION>"
        else:
            task_prompt = "<DETAILED_CAPTION>"

        config = self.get_generation_config(prompt_mode)
        inputs = self.processor(text=task_prompt, images=image, return_tensors="pt").to(self.device)
        try:
            with torch.no_grad():
                out_ids = self.model.generate(
                    input_ids=inputs["input_ids"],
                    pixel_values=inputs["pixel_values"],
                    **config
                )
            gen_text = self.processor.batch_decode(out_ids, skip_special_tokens=False)[0]
            result = self.processor.post_process_generation(
                gen_text,
                image_size=image.size,
                task=task_prompt
            ).get(task_prompt, "")

            if max_words:
                result = " ".join(result.split()[:max_words])
            return f"{prefix.strip()} {result} {suffix.strip()}".strip()
        except Exception as e:
            logger.error(f"[Florence-2 Error] {e}")
            return "an image"


class CLIPPromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device):
        super().__init__(device)
        from clip_interrogator import Config, Interrogator
        self.config = Config(
            clip_model_name="ViT-L-14/openai",
            device="cuda" if device.type == "cuda" else "cpu"
        )
        self.interrogator = Interrogator(self.config)
        self.prompt_cache = {}

    def generate_prompt(
        self,
        image: Image.Image,
        prompt_mode: str = "Complex",
        max_words: Optional[int] = None,
        prefix: str = "",
        suffix: str = ""
    ) -> str:
        img_hash = self._hash(image)
        if img_hash in self.prompt_cache:
            return self.prompt_cache[img_hash]

        try:
            if prompt_mode.lower() == "simple":
                caption = self.interrogator.interrogate_fast(image)
            else:
                caption = self.interrogator.interrogate(image)

            if max_words:
                caption = " ".join(caption.split()[:max_words])

            out = f"{prefix.strip()} {caption} {suffix.strip()}".strip()
            self.prompt_cache[img_hash] = out
            return out
        except Exception as e:
            logger.error(f"[CLIP Error] {e}")
            return "an image"

    def _hash(self, image: Image.Image) -> str:
        buf = io.BytesIO()
        image.save(buf, format="PNG")
        import hashlib
        return hashlib.md5(buf.getvalue()).hexdigest()


# -------------------------------------------------
# USDURedraw
# -------------------------------------------------
TILE_SIZE_MIN = 1024
TILE_SIZE_MAX = 1536
TILE_SIZE_STEP = 64
DEFAULT_TILE_SIZE = 1024
DEFAULT_FEATHER = 128
DEFAULT_OVERLAP = 20

class USDURedraw:
    """Handles tiling, prompt gen, filtering, reassembly."""
    def __init__(
        self,
        usdupscaler,
        device,
        feather_amount: int = 50,
        save_tiles: bool = False,
        turn_to_night: bool = False,
        enable_content_analysis: bool = False,
        min_keyword_overlap: int = 2,
        context_window_scale: float = 2.0,
        prompt_mode_tile: str = "Simple",
        exclude_categories: Dict[str, bool] = None,
        # Additional user-provided words to remove
        user_words_to_remove: Optional[str] = ""
    ):
        self.upscaler = usdupscaler
        self.device = device
        self.feather_amount = feather_amount
        self.save_tiles = save_tiles
        self.tile_size = 1024
        self.base_prompt = ""
        self.clip_prompt_suffix = ""
        self.entire_image_prompt = ""

        self.prompt_generator = usdupscaler.prompt_generator
        self.turn_to_night = turn_to_night and NIGHT_MODE_AVAILABLE
        self.enable_content_analysis = enable_content_analysis
        self.min_keyword_overlap = min_keyword_overlap
        self.context_window_scale = context_window_scale
        self.prompt_mode_tile = prompt_mode_tile
        self.save = False

        # Read the 8 categories from categories.yaml
        self.category_words = read_categories_yaml()
        if not self.category_words:
            logger.warning("No categories found in categories.yaml - filters won't do anything.")
            self.category_words = {}

        # Which categories are excluded?
        self.exclude_categories = exclude_categories or {}

        # Additional user words
        self.words_to_remove: List[str] = []
        if user_words_to_remove.strip():
            # Split by comma
            self.words_to_remove = [w.strip().lower() for w in user_words_to_remove.split(",") if w.strip()]

        # DOF
        self.dof_handler = DepthOfFieldFeatures()

        # Tiling
        self.cols = 1
        self.rows = 1
        self.width = 0
        self.height = 0
        self.tile_positions_x = []
        self.tile_positions_y = []
        self.overlap_percentage = 0
        self.overlap = 0

    def analyze_full_image(self):
        if not self.prompt_generator:
            self.entire_image_prompt = self.upscaler.p.prompt or ""
            return

        self.entire_image_prompt = self.prompt_generator.generate_prompt(
            self.upscaler.image,
            prompt_mode="Complex",
            prefix=self.base_prompt,
            suffix=self.clip_prompt_suffix
        )
        self.entire_image_prompt = self._apply_content_filters(self.entire_image_prompt)

    def generate_tile_prompt(self, tile_image: Image.Image) -> str:
        if not self.prompt_generator:
            return self.upscaler.p.prompt

        tile_prompt = self.prompt_generator.generate_prompt(
            tile_image,
            prompt_mode=self.prompt_mode_tile,
            prefix=self.base_prompt,
            suffix=self.clip_prompt_suffix
        )
        logger.debug(f"Raw generated prompt [{self.prompt_mode_tile}]: {tile_prompt}")

        tile_prompt = self._apply_content_filters(tile_prompt)
        logger.debug(f"After content filters => {tile_prompt}")

        if self.enable_content_analysis:
            tile_prompt = self._analyze_and_enhance_prompt(tile_image, tile_prompt)
            logger.debug(f"After content analysis => {tile_prompt}")

        return tile_prompt

    def _apply_content_filters(self, prompt: str) -> str:
        if not prompt:
            return prompt

        # 1) Exclude categories that are checked
        for cat_name, do_exclude in self.exclude_categories.items():
            if do_exclude and cat_name in self.category_words:
                prompt = self._remove_terms(prompt, self.category_words[cat_name])

        # 2) Remove user-specified words
        if self.words_to_remove:
            prompt = self._remove_terms(prompt, self.words_to_remove)

        # 3) Night mode
        if self.turn_to_night:
            try:
                new_prompt = convert_to_night_prompt(prompt)
                if new_prompt != prompt:
                    logger.info(f"Night mode changed prompt from: '{prompt}' to '{new_prompt}'")
                prompt = new_prompt
            except Exception as e:
                logger.error(f"Night mode error: {e}")

        return prompt.strip()

    def _remove_terms(self, prompt: str, terms: List[str]) -> str:
        if not prompt or not terms:
            return prompt
        pattern = r'\b(' + '|'.join(map(re.escape, terms)) + r')(s|es|\'s)?\b'
        cleaned = re.sub(pattern, '', prompt, flags=re.IGNORECASE)
        return ' '.join(cleaned.split())

    def _analyze_and_enhance_prompt(self, image: Image.Image, prompt: str) -> str:
        current_keywords = set(self.get_important_keywords(prompt))
        global_keywords = set(self.get_important_keywords(self.entire_image_prompt))

        try:
            depth_score, _ = self.dof_handler.analyze_region_depth(image)
            enhanced_prompt = self.dof_handler.analyze_dof_context(
                prompt,
                context_keywords=current_keywords,
                global_keywords=global_keywords,
                depth_value=depth_score
            )
            if len(current_keywords & global_keywords) < self.min_keyword_overlap:
                enhanced_prompt += " Maintain consistency with overall image."
            return enhanced_prompt
        except Exception as e:
            logger.error(f"Error in content analysis: {e}")
            return prompt

    def get_important_keywords(self, text: str) -> List[str]:
        common_words = {
            'a','an','the','in','on','at','with','and','or','of','to','for',
            'is','are','was','were','be','been','being','by','that','this',
            'these','those','it','its','from','as','has','have','had'
        }
        if not text:
            return []
        words = text.lower().split()
        return [w for w in words if w and len(w) > 1 and w not in common_words]

    # -------------------------------------------
    # Multi-tile
    # -------------------------------------------
    def pre_analyze(self, image: Image.Image) -> List[List[str]]:
        self.analyze_full_image()
        global_keywords = set(self.get_important_keywords(self.entire_image_prompt))
        logger.info(f"[Global Keywords] => {global_keywords}")
        return self._generate_tile_prompts(image)

    def _generate_tile_prompts(self, image: Image.Image) -> List[List[str]]:
        tile_prompts = []
        for yi in range(self.rows):
            row_prompts = []
            for xi in range(self.cols):
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)
                if x2 <= x1 or y2 <= y1:
                    row_prompts.append("")
                    continue
                tile_img = image.crop((x1, y1, x2, y2))
                tile_prompt = self.generate_tile_prompt(tile_img)
                row_prompts.append(tile_prompt)
            tile_prompts.append(row_prompts)
        return tile_prompts

    def calc_rectangle(self, xi: int, yi: int) -> Tuple[int,int,int,int]:
        x1 = self.tile_positions_x[xi]
        y1 = self.tile_positions_y[yi]
        x2 = x1 + self.tile_size if xi < self.cols - 1 else self.width
        y2 = y1 + self.tile_size if yi < self.rows - 1 else self.height
        return x1, y1, x2, y2

    # -------------------------------------------
    # Final Tiled Processing
    # -------------------------------------------
    def start(self, p, image: Image.Image, tile_prompts: List[List[str]]) -> Image.Image:
        self.width, self.height = image.size
        final_image = Image.new("RGB", (self.width, self.height))

        for yi in range(self.rows):
            for xi in range(self.cols):
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)
                if x2 <= x1 or y2 <= y1:
                    continue

                tile_img = image.crop((x1, y1, x2, y2))
                tile_prompt = tile_prompts[yi][xi]
                logger.info(f"[Tile({xi},{yi})] => {tile_prompt}")

                tile_p = self.prepare_tile_processing(p, tile_img, tile_prompt)
                processed = processing.process_images(tile_p)
                processed_tile = processed.images[0].resize(tile_img.size, Image.LANCZOS)

                if self.save_tiles:
                    tile_fname = f"tile_{yi}_{xi}.png"
                    tile_path = os.path.join(p.outpath_samples, tile_fname)
                    processed_tile.save(tile_path)
                    logger.info(f"Saved tile => {tile_path}")

                # Feather edges
                if self.feather_amount > 0:
                    mask = self.create_feathered_mask((x1, y1, x2, y2), (self.width, self.height), self.feather_amount)
                    if mask.size != tile_img.size:
                        mask = mask.resize(tile_img.size, Image.LANCZOS)
                else:
                    mask = Image.new("L", tile_img.size, 255)

                final_image.paste(processed_tile, (x1, y1), mask)

                if self.device.type == 'cuda':
                    torch.cuda.empty_cache()
                del tile_img, processed, processed_tile, mask, tile_p
                gc.collect()

                state.job_no += 1
                if state.interrupted:
                    break
            if state.interrupted:
                break

        return final_image

    def prepare_tile_processing(self, p, tile_img: Image.Image, tile_prompt: str):
        tile_p = copy.copy(p)
        tile_p.width = min(tile_img.width, max(64, (tile_img.width // 64) * 64))
        tile_p.height = min(tile_img.height, max(64, (tile_img.height // 64) * 64))
        resized = tile_img.resize((tile_p.width, tile_p.height), Image.LANCZOS)

        tile_p.init_images = [resized]
        tile_p.prompt = tile_prompt
        tile_p.batch_size = 1
        tile_p.n_iter = 1
        tile_p.do_not_save_samples = not self.save_tiles
        tile_p.do_not_save_grid = True
        fix_seed(tile_p)
        return tile_p

    def create_feathered_mask(self, tile_region: Tuple[int,int,int,int],
                              img_size: Tuple[int,int],
                              radius: int) -> Image.Image:
        tile_w = tile_region[2] - tile_region[0]
        tile_h = tile_region[3] - tile_region[1]
        x1, y1, x2, y2 = tile_region
        w, h = img_size

        alpha = Image.new("L", (tile_w, tile_h), 255)
        d = ImageDraw.Draw(alpha)

        # Edges
        feather_left = x1 > 0
        feather_top = y1 > 0
        feather_right = x2 < w
        feather_bottom = y2 < h

        def s_curve(xf: float) -> int:
            xf = xf * 8.0 - 4.0
            yv = 1 / (1 + math.exp(-xf))
            yv = math.pow(yv, 2.0)
            return int(yv * 255)

        # Feather lines
        if feather_left:
            for i in range(radius):
                opacity = s_curve(i / radius)
                d.line([(i,0),(i,tile_h)], fill=opacity, width=1)
        if feather_right:
            for i in range(radius):
                opacity = s_curve(i / radius)
                d.line([(tile_w - i - 1,0),(tile_w - i - 1,tile_h)], fill=opacity, width=1)
        if feather_top:
            for i in range(radius):
                opacity = s_curve(i / radius)
                d.line([(0,i),(tile_w,i)], fill=opacity, width=1)
        if feather_bottom:
            for i in range(radius):
                opacity = s_curve(i / radius)
                d.line([(0,tile_h - i - 1),(tile_w,tile_h - i - 1)], fill=opacity, width=1)

        mask = alpha.filter(ImageFilter.GaussianBlur(radius=radius // 4))
        return mask


# -------------------------------------------------
# USDUpscaler
# -------------------------------------------------
class USDUpscaler:
    model_cache = {}

    def __init__(
        self,
        p,
        image: Image.Image,
        upscaler_index: int,
        save_redraw: bool,
        tile_size: int,
        tile_size_min: int,
        tile_size_max: int,
        base_prompt: str,
        overlap_percentage: float,
        clip_prompt_suffix: str,
        prompt_method: str,
        prompt_mode_tile: str,
        save_prompts_to_file: bool,
        uploaded_prompt_file,
        global_seed: int,
        feather_amount: int,
        save_tiles: bool,
        turn_to_night: bool,
        enable_content_analysis: bool,
        min_keyword_overlap: int,
        context_window_scale: float,
        exclude_categories: Dict[str, bool],
        user_words_to_remove: str
    ):
        self.p = p
        self.image = image
        self.prompt_mode_tile = prompt_mode_tile
        self.global_seed = global_seed
        self.save_prompts_to_file = save_prompts_to_file
        self.uploaded_prompt_file = uploaded_prompt_file
        self.result_images = []
        self.job_count = 0

        init_w, init_h = image.size
        scale_w = p.width / init_w
        scale_h = p.height / init_h
        if abs(scale_w - scale_h) > 0.01:
            logger.warning("Aspect ratio changed. Using width-based scale.")
        self.scale_factor = scale_w

        # Match upscaler
        self.upscaler_data = (
            shared.sd_upscalers[upscaler_index]
            if 0 <= upscaler_index < len(shared.sd_upscalers)
            else None
        )
        self.prompt_generator = self.init_generator(prompt_method)

        self.redraw = USDURedraw(
            usdupscaler=self,
            device=torch.device("cuda" if torch.cuda.is_available() else "cpu"),
            feather_amount=feather_amount,
            save_tiles=save_tiles,
            turn_to_night=turn_to_night,
            enable_content_analysis=enable_content_analysis,
            min_keyword_overlap=min_keyword_overlap,
            context_window_scale=context_window_scale,
            prompt_mode_tile=prompt_mode_tile,
            exclude_categories=exclude_categories,
            user_words_to_remove=user_words_to_remove
        )
        self.redraw.base_prompt = base_prompt
        self.redraw.clip_prompt_suffix = clip_prompt_suffix
        self.redraw.save = save_redraw

        # (1) Upscale entire image if chosen
        self.upscale_image()
        # (2) Setup tiles
        self.setup_tiles(tile_size, overlap_percentage, tile_size_min, tile_size_max)

    def init_generator(self, method: str):
        if method == "NONE":
            return None
        if method in USDUpscaler.model_cache:
            return USDUpscaler.model_cache[method]

        dev = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        try:
            if method == "BLIP":
                gen = BLIPPromptGenerator(dev)
            elif method == "ViT-GPT2":
                gen = ViTGPT2PromptGenerator(dev)
            elif method == "Florence-2":
                gen = Florence2PromptGenerator(dev, task_option="Detailed Caption")
            elif method == "CLIP":
                gen = CLIPPromptGenerator(dev)
            else:
                raise ValueError(f"Unsupported prompt method: {method}")

            USDUpscaler.model_cache[method] = gen
            return gen
        except Exception as e:
            logger.error(f"[PromptGen Error] {e}")
            return None

    def upscale_image(self):
        if self.upscaler_data:
            logger.info(f"Upscaling with {self.upscaler_data.name} ...")
            self.image = self.upscaler_data.scaler.upscale(
                self.image,
                self.scale_factor,
                self.upscaler_data.data_path
            )
            self.p.width, self.p.height = self.image.size
            logger.info(f"Upscaled => {self.image.size}")
        else:
            logger.info("No upscaler selected. Skipping upscaling step.")

    def setup_tiles(
        self,
        tile_size: int,
        overlap_percentage: float,
        tile_size_min: int,
        tile_size_max: int
    ):
        w, h = self.image.size

        if tile_size_min < tile_size_max:
            best_tile_size = tile_size_min
            min_tile_count = float('inf')

            for test_size in range(tile_size_min, tile_size_max + 1, TILE_SIZE_STEP):
                test_overlap = int(overlap_percentage / 100 * test_size)
                tcols = max(1, math.ceil((w + test_overlap) / (test_size - test_overlap)))
                trows = max(1, math.ceil((h + test_overlap) / (test_size - test_overlap)))
                total_tiles = tcols * trows
                if total_tiles < min_tile_count:
                    min_tile_count = total_tiles
                    best_tile_size = test_size
                elif total_tiles > min_tile_count:
                    break

            self.redraw.tile_size = best_tile_size
            logger.info(f"Auto-selected tile size => {best_tile_size} ({min_tile_count} tiles).")
        else:
            bounded_size = max(tile_size_min, min(tile_size_max, tile_size))
            self.redraw.tile_size = bounded_size
            logger.info(f"Using tile_size => {bounded_size}")

        self.redraw.overlap_percentage = overlap_percentage
        self.redraw.overlap = int(overlap_percentage / 100 * self.redraw.tile_size)

        self.redraw.cols = max(
            1, math.ceil((w + self.redraw.overlap) / (self.redraw.tile_size - self.redraw.overlap))
        )
        self.redraw.rows = max(
            1, math.ceil((h + self.redraw.overlap) / (self.redraw.tile_size - self.redraw.overlap))
        )
        self.redraw.width = w
        self.redraw.height = h

        self.redraw.tile_positions_x = []
        for i in range(self.redraw.cols):
            if i == self.redraw.cols - 1:
                xx = w - self.redraw.tile_size
            else:
                xx = i * (self.redraw.tile_size - self.redraw.overlap)
            self.redraw.tile_positions_x.append(xx)

        self.redraw.tile_positions_y = []
        for j in range(self.redraw.rows):
            if j == self.redraw.rows - 1:
                yy = h - self.redraw.tile_size
            else:
                yy = j * (self.redraw.tile_size - self.redraw.overlap)
            self.redraw.tile_positions_y.append(yy)

        logger.info(f"Grid => {self.redraw.cols} x {self.redraw.rows}")

    def pre_analyze_tiles(self):
        if self.uploaded_prompt_file:
            self.tile_prompts = self.load_excel(self.uploaded_prompt_file)
        else:
            self.tile_prompts = self.redraw.pre_analyze(self.image)
            if self.save_prompts_to_file and self.prompt_generator:
                self.save_prompts_xlsx(self.tile_prompts)

    def load_excel(self, upf):
        wb = load_workbook(filename=upf.name)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            row_vals = []
            for c in row:
                row_vals.append(c if c else "")
            data.append(row_vals)

        if len(data) != self.redraw.rows or max(len(r) for r in data) != self.redraw.cols:
            raise ValueError("[Excel mismatch] The file's rows/cols do not match tile grid.")
        return data

    def process(self):
        state.begin()
        self.calc_jobs()
        self.pre_analyze_tiles()
        out_img = self.redraw.start(self.p, self.image, self.tile_prompts)
        self.result_images.append(out_img)

        if self.redraw.save:
            saved_path = self.save_final_image(out_img)
            if isinstance(saved_path, str) and self.save_prompts_to_file and not self.uploaded_prompt_file:
                self.save_prompts_xlsx(self.tile_prompts, final_image_path=saved_path)

        self.cleanup()
        state.end()

    def calc_jobs(self):
        self.job_count = self.redraw.cols * self.redraw.rows
        state.job_count = self.job_count

    def save_final_image(self, outimg: Image.Image):
        from modules import images

        method_name = self.prompt_generator.__class__.__name__ if self.prompt_generator else "NONE"
        method_suffix = f"_{self.prompt_mode_tile}_{method_name}"
        metadata = {
            "UpscalerName": self.upscaler_data.name if self.upscaler_data else "None",
            "TileSize": str(self.redraw.tile_size),
            "OverlapPercentage": str(self.redraw.overlap_percentage),
            "FeatherAmount": str(self.redraw.feather_amount),
            "PromptMethod": method_name,
            "PromptMode": str(self.prompt_mode_tile),
            "BasePrompt": str(self.redraw.base_prompt),
            "Seed": str(self.global_seed),
            "GridSize": f"{self.redraw.cols}x{self.redraw.rows}",
            "FinalSize": f"{self.redraw.width}x{self.redraw.height}"
        }
        info = "Upscaling parameters:\n" + "\n".join(f"{k}: {v}" for k,v in metadata.items())

        fpath = images.save_image(
            outimg,
            self.p.outpath_samples,
            "",
            seed=self.global_seed,
            prompt=self.p.prompt,
            extension=opts.samples_format,
            info=info,
            p=self.p,
            suffix=method_suffix,
            pnginfo_section_name="RegionalUpscaler"
        )
        logger.info(f"Final image => {fpath}")
        return fpath

    def save_prompts_xlsx(self, tile_prompts: List[List[str]], final_image_path: Optional[str] = None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tile Prompts"

        for y, rowprm in enumerate(tile_prompts):
            for x, prm in enumerate(rowprm):
                c = ws.cell(row=y + 1, column=x + 1)
                c.value = prm
                c.alignment = Alignment(wrap_text=True)

        for col in ws.columns:
            cells = list(col)
            if not cells:
                continue
            length = max(len(str(cell.value)) for cell in cells if cell.value) if cells else 0
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, 50)

        method_name = self.prompt_generator.__class__.__name__ if self.prompt_generator else "NONE"
        output_excel_path = os.path.join(self.p.outpath_samples, f"tile_prompts_{method_name}.xlsx")
        attempt = 1
        while True:
            try:
                wb.save(output_excel_path)
                logger.info(f"Excel saved => {output_excel_path}")
                break
            except PermissionError:
                output_excel_path = f"{os.path.splitext(output_excel_path)[0]}_{attempt}.xlsx"
                attempt += 1

        if final_image_path:
            base, ext = os.path.splitext(final_image_path)
            image_excel_path = f"{base}_prompts.xlsx"
            attempt = 1
            while True:
                try:
                    wb.save(image_excel_path)
                    logger.info(f"Excel saved alongside image => {image_excel_path}")
                    break
                except PermissionError:
                    image_excel_path = f"{os.path.splitext(image_excel_path)[0]}_{attempt}.xlsx"
                    attempt += 1

    def cleanup(self):
        if self.prompt_generator:
            del self.prompt_generator
        torch.cuda.empty_cache()
        gc.collect()


# -------------------------------------------------
# The A1111 Script
# -------------------------------------------------
class Script(scripts.Script):
    # We only have 8 categories to show, in exact order:
    # "humans", "animals", "vehicles", "sky", "water", "tech", "food", "plants"
    # (Vehicles labeled "Vehicles (except boats)")
    category_order = [
        ("humans", "Exclude Humans"),
        ("animals", "Exclude Animals"),
        ("vehicles", "Exclude Vehicles (except boats)"),
        ("sky", "Exclude Sky"),
        ("water", "Exclude Water"),
        ("tech", "Exclude Tech"),
        ("food", "Exclude Food"),
        ("plants", "Exclude Plants"),
    ]

    def title(self):
        return "Regional Prompt Upscaler and Detailer v1.3"

    def show(self, is_img2img):
        return is_img2img

    def get_lora_files(self):
        lora_path = os.path.join("models", "Lora")
        if not os.path.exists(lora_path):
            return ["None"]
        all_loras = []
        for root, dirs, files in os.walk(lora_path):
            for f in files:
                if f.lower().endswith(".safetensors") or f.lower().endswith(".ckpt"):
                    rp = os.path.relpath(os.path.join(root, f), lora_path)
                    rp = rp.replace(os.path.sep, "/")
                    all_loras.append(f"/LoRA/{rp}")
        all_loras.sort()
        return ["None"] + all_loras

    def ui(self, is_img2img):
        with gr.Blocks():
            with gr.Row():
                gr.HTML("<h3>Regional Prompt Upscaler and Detailer v1.3</h3>")
                gr.Markdown("Tile-based upscaling with advanced prompt injection + category filtering.")

            with gr.Row():
                with gr.Column():
                    with gr.Box():
                        gr.HTML("<b>Upscaler Settings</b>")
                        upscaler_name = gr.Dropdown(
                            label='Upscaler Model',
                            choices=[x.name for x in shared.sd_upscalers],
                            value=(shared.sd_upscalers[0].name if len(shared.sd_upscalers) else "None")
                        )
                        with gr.Row():
                            tile_size_min = gr.Slider(label='Min Tile Size', minimum=1024, maximum=1536, step=64, value=1024)
                            tile_size_max = gr.Slider(label='Max Tile Size', minimum=1024, maximum=4096, step=64, value=1536)
                        tile_size = gr.Slider(
                            label='Preferred Tile Size',
                            minimum=1024, maximum=2048, step=64, value=1024
                        )
                        overlap_percentage = gr.Slider(label='Tile Overlap (%)', minimum=0, maximum=50, step=1, value=20)
                        feather_amount = gr.Slider(label='Feather Amount', minimum=0, maximum=384, step=1, value=128)

                with gr.Column():
                    with gr.Box():
                        gr.HTML("<b>Prompt Generation</b>")
                        prompt_method = gr.Dropdown(
                            label="Method",
                            choices=["BLIP", "ViT-GPT2", "Florence-2", "CLIP", "NONE"],
                            value="Florence-2"
                        )
                        prompt_mode_tile = gr.Dropdown(
                            label="Prompt Complexity",
                            choices=["Simple", "Complex"],
                            value="Simple"
                        )
                        base_prompt = gr.Textbox(label="Prefix", lines=1, value="")
                        clip_prompt_suffix = gr.Textbox(label="Suffix", lines=1, value="")
                        with gr.Row():
                            lora_model = gr.Dropdown(
                                label="LoRA",
                                choices=self.get_lora_files(),
                                value="None",
                                scale=10
                            )
                            refresh_lora_btn = gr.Button(value="â†»", variant="secondary", size="sm", min_width=10)

            with gr.Row():
                with gr.Column():
                    with gr.Box():
                        gr.HTML("<b>Filtering & Night Mode</b>")
                        turn_to_night = gr.Checkbox(label="Night Mode - Requires Add-on From Hallett-ai.com", value=False)

                        # The 8 categories, in specified order
                        cat_checkboxes = {}
                        for cat_key, label_text in self.category_order:
                            cat_checkboxes[cat_key] = gr.Checkbox(label=label_text, value=False)

                        words_to_remove = gr.Textbox(label="Remove Words (comma-separated)", lines=1, value="")
                with gr.Column():
                    with gr.Box():
                        gr.HTML("<b>Other Settings</b>")
                        save_tiles = gr.Checkbox(label="Save Individual Tiles", value=False)
                        save_upscaled_image = gr.Checkbox(label="Save Final Image", value=True)
                        save_prompts_to_file = gr.Checkbox(label="Save Tile Prompts (Excel)", value=True)
                        upload_prompt_file = gr.File(label="Import Tile Prompts (.xlsx)", file_types=[".xlsx"])
                        debug_logging = gr.Checkbox(label="Enable Debug Logging", value=False)

            gr.HTML("<hr style='margin: 15px 0; border: none; border-top: 1px solid rgba(128,128,128,0.2);'>")
            enable_content_analysis = gr.Checkbox(label='Enable Content Analysis', value=False)
            min_keyword_overlap = gr.Slider(
                label="Keyword Overlap (1-5)",
                minimum=1, maximum=5, step=1, value=2, visible=False
            )
            context_window_scale = gr.Slider(
                label="Context Window Scale (1.5 - 3.0)",
                minimum=1.5, maximum=3.0, step=0.5, value=2.0, visible=False
            )

            # Show/hide advanced content sliders
            enable_content_analysis.change(
                lambda x: [gr.update(visible=x), gr.update(visible=x)],
                inputs=[enable_content_analysis],
                outputs=[min_keyword_overlap, context_window_scale]
            )

            # Refresh lora
            refresh_lora_btn.click(
                lambda: gr.update(choices=self.get_lora_files()),
                inputs=[],
                outputs=[lora_model]
            )

            # If user picks a LoRA, add prefix
            def set_lora_prefix(sel):
                return set_lora_in_prefix(sel)

            lora_model.change(
                set_lora_prefix,
                inputs=[lora_model],
                outputs=[base_prompt]
            )

            def update_debug_logging(val):
                global logger
                logger = setup_logging(debug_enabled=val)
                return "Debug logging updated."

            debug_logging.change(update_debug_logging, inputs=[debug_logging], outputs=[])

            # Return all 21 parameters in order
            # First 20 known args + category checkboxes
            # Important: This must match exactly what run() expects
            return [
                enable_content_analysis,
                upscaler_name,
                tile_size, tile_size_min, tile_size_max,
                feather_amount, overlap_percentage,
                min_keyword_overlap, context_window_scale,
                base_prompt, clip_prompt_suffix,
                prompt_method, prompt_mode_tile,
                save_prompts_to_file, upload_prompt_file,
                save_upscaled_image, save_tiles,
                turn_to_night,
                lora_model,
                debug_logging,
                words_to_remove
            ] + [cat_checkboxes[key] for key, _ in self.category_order]  # Ensure ordered checkboxes

    def run(self, p, *all_args):
        """
        The order of arguments is exactly as returned in ui():
        - 20 known arguments up to words_to_remove
        - + 8 category checkboxes in the specified order

        We'll parse them carefully below.
        """
        # Split arguments into known args and category values
        known_args = all_args[:21]  # Ensure we get all 21 parameters
        cat_checkbox_values = all_args[21:]  # 8 category checkboxes

        # Unpack known arguments in the same order as ui() returns them
        (
            enable_content_analysis,
            upscaler_name,
            tile_size, tile_size_min, tile_size_max,
            feather_amount, overlap_percentage,
            min_keyword_overlap, context_window_scale,
            base_prompt, clip_prompt_suffix,
            prompt_method, prompt_mode_tile,
            save_prompts_to_file, upload_prompt_file,
            save_upscaled_image, save_tiles,
            turn_to_night,
            lora_model_name,
            debug_logging,
            words_to_remove
        ) = known_args

        # Update logging if needed
        global logger
        logger = setup_logging(debug_enabled=debug_logging)

        # Build exclude_categories from the 8 checkboxes
        exclude_categories = {}
        for i, (cat_key, _) in enumerate(self.category_order):
            exclude_categories[cat_key] = bool(cat_checkbox_values[i])

        if not enable_content_analysis:
            min_keyword_overlap = 2
            context_window_scale = 2.0

        fix_seed(p)
        global_seed = p.seed
        init_img = p.init_images[0] if p.init_images else None
        if not init_img:
            return Processed(p, [], p.seed, "[ERROR] No initial image provided.")

        init_img = images.flatten(init_img, opts.img2img_background_color)

        # upscaler name => index
        upscaler_index = 0
        for i, u in enumerate(shared.sd_upscalers):
            if u.name == upscaler_name:
                upscaler_index = i
                break

        # Orchestrator
        upscaler = USDUpscaler(
            p=p,
            image=init_img,
            upscaler_index=upscaler_index,
            save_redraw=save_upscaled_image,
            tile_size=tile_size,
            tile_size_min=tile_size_min,
            tile_size_max=tile_size_max,
            base_prompt=base_prompt,
            overlap_percentage=overlap_percentage,
            clip_prompt_suffix=clip_prompt_suffix,
            prompt_method=prompt_method,
            prompt_mode_tile=prompt_mode_tile,
            save_prompts_to_file=save_prompts_to_file,
            uploaded_prompt_file=upload_prompt_file,
            global_seed=global_seed,
            feather_amount=feather_amount,
            save_tiles=save_tiles,
            turn_to_night=turn_to_night,
            enable_content_analysis=enable_content_analysis,
            min_keyword_overlap=min_keyword_overlap,
            context_window_scale=context_window_scale,
            exclude_categories=exclude_categories,
            user_words_to_remove=words_to_remove
        )

        # Execute
        upscaler.process()
        return Processed(p, upscaler.result_images, p.seed, "Done.")
