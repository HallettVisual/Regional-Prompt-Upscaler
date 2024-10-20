import subprocess
import sys
import importlib

# Utility functions
# Function to install a Python package
def install(package):
    try:
        print(f"Installing package: {package}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    except Exception as e:
        print(f"Failed to install {package}: {e}")

# Function to ensure a package is installed
def ensure_package_installed(package, module_name=None):
    if module_name is None:
        module_name = package
    try:
        __import__(module_name)
    except ImportError:
        print(f"{package} is not installed. Attempting to install...")
        install(package)

# Ensure spacy is installed
ensure_package_installed('spacy')

# Now that spacy is installed, we import it
import spacy  # Natural Language Processing

# Ensure the SpaCy model 'en_core_web_sm' is installed
def ensure_spacy_model_installed(model):
    try:
        spacy.load(model)
    except OSError:
        print(f"Spacy model '{model}' not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "spacy", "download", model])
        print(f"Spacy model '{model}' installed successfully.")

# Now ensure the SpaCy model is installed
ensure_spacy_model_installed('en_core_web_sm')

# Now that dependencies are ensured, import other necessary modules
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Standard library imports and core dependencies
import os
import math
import re  # For regular expressions
import copy  # For making deep copies of objects
from enum import Enum

# External libraries
import cv2  # OpenCV for blurriness detection
import torch  # PyTorch
import numpy as np
from PIL import Image, ImageDraw, ImageFilter
from transformers import (
    BlipProcessor, BlipForConditionalGeneration,
    CLIPProcessor, CLIPModel
)
import gradio as gr  # Gradio UI library

# Stable Diffusion WebUI-specific imports
from modules import processing, shared, images, devices, scripts
from modules.processing import StableDiffusionProcessing, Processed, fix_seed
from modules.shared import opts, state

# Initialize device
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

elem_id_prefix = "promptupscaler"

# Base class for Prompt Generators
class PromptGeneratorBase:
    def __init__(self, device: torch.device):
        self.device = device

    def generate_prompt(self, image: Image.Image, max_words: int, prefix: str = "", suffix: str = "") -> str:
        raise NotImplementedError("generate_prompt method must be implemented by subclasses.")

    def supports_save_load(self) -> bool:
        """Indicates whether this generator supports prompt saving and loading."""
        return True

# BLIP Prompt Generator
class BLIPPromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device):
        super().__init__(device)
        self.processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base")
        self.model = BlipForConditionalGeneration.from_pretrained("Salesforce/blip-image-captioning-base").to(self.device)
        self.model.eval()
        torch.cuda.empty_cache()

    def generate_prompt(self, image: Image.Image, max_words: int, prefix: str = "", suffix: str = "") -> str:
        try:
            inputs = self.processor(images=image, return_tensors="pt").to(self.device)
            with torch.no_grad():
                caption_ids = self.model.generate(
                    **inputs,
                    max_length=max_words,
                    min_length=int(max_words * 0.8),
                    no_repeat_ngram_size=2,
                    length_penalty=1.0,
                    num_beams=5,
                    early_stopping=True
                )
            caption = self.processor.decode(caption_ids[0], skip_special_tokens=True)
            prompt = f"{prefix} {caption} {suffix}".strip()
            return prompt
        finally:
            del inputs
            del caption_ids
            torch.cuda.empty_cache()

# CLIP Prompt Generator
class CLIPPromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device):
        super().__init__(device)
        self.model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32").to(self.device)
        self.processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")
        self.model.eval()
        torch.cuda.empty_cache()
        # Load labels (e.g., ImageNet class names or a custom list)
        self.labels = self.load_labels()

    def load_labels(self):
        # Load a comprehensive list of labels
        return [
            # Add a comprehensive list of labels relevant to your images
            "a dog", "a cat", "a person", "a landscape", "a building", "a car",
            "a tree", "a river", "a mountain", "a beach", "a sunset", "a forest",
            "an ocean", "a street", "a flower", "an animal", "a bird", "the sky",
            "a cityscape", "a boat", "a night scene", "snow", "rain", "a desert",
            "clouds", "a bridge", "fireworks", "a garden", "a park", "a waterfall",
            # Add more labels as needed
        ]

    def generate_prompt(self, image: Image.Image, max_words: int, prefix: str = "", suffix: str = "") -> str:
        try:
            # Prepare inputs using the labels
            inputs = self.processor(text=self.labels, images=image, return_tensors="pt", padding=True).to(self.device)

            # Get the similarity scores between image and labels
            with torch.no_grad():
                outputs = self.model(**inputs)
                logits_per_image = outputs.logits_per_image  # (1, len(labels))
                probs = logits_per_image.softmax(dim=1)  # shape: (1, len(labels))

            # Select top labels that match the image
            top_k = min(max_words, len(self.labels))  # Number of labels to select
            top_probs, top_idxs = probs[0].topk(k=top_k)
            top_labels = [self.labels[idx] for idx in top_idxs]

            # Remove duplicates and clean up labels
            top_labels = list(dict.fromkeys(top_labels))  # Remove duplicates while preserving order

            # Construct the prompt
            prompt_components = [prefix.strip()] + top_labels + [suffix.strip()]
            prompt = ', '.join(filter(None, prompt_components))  # Use commas to separate labels
            prompt_words = prompt.split()[:max_words]  # Truncate to max_words

            return ' '.join(prompt_words)
        finally:
            # Clean up to prevent memory leaks
            del inputs
            torch.cuda.empty_cache()

# USDURedraw Class
class USDURedraw:
    def __init__(
        self,
        usdupscaler,
        device,
        master_prompt_max_words=75,
        tile_prompt_max_words=10,
        feather_amount=50,
        blur_threshold=1,
        low_detail_threshold=1,
        save_tiles=False,
        prompt_generator=None,
        words_to_remove=None,
        categories_to_remove=None,  # Added categories to remove
    ):
        self.upscaler = usdupscaler
        self.device = device
        self.save = False
        self.save_tiles = save_tiles
        self.tile_size = 1024
        self.enabled = True
        self.max_clip_words = tile_prompt_max_words
        self.master_prompt_max_words = master_prompt_max_words
        self.base_prompt = ""
        self.initial_info = None
        self.overlap_percentage = 0
        self.clip_prompt_suffix = ""
        self.entire_image_prompt = ""
        self.feather_amount = feather_amount
        self.prompt_generator = prompt_generator
        self.words_to_remove = words_to_remove if words_to_remove else []
        self.categories_to_remove = categories_to_remove if categories_to_remove else []
        self.category_words = {
            'animals': ['animal', 'dog', 'cat', 'bird', 'lion', 'tiger', 'elephant', 'horse', 'fish', 'insect', 'bear', 'wolf', 'rabbit', 'fox'],
            'sky': ['sky', 'cloud', 'plane', 'bird', 'sun', 'moon', 'star', 'rainbow', 'aeroplane', 'helicopter', 'aircraft', 'sunset', 'sunrise'],
            'water': ['water', 'lake', 'river', 'ocean', 'sea', 'pond', 'stream', 'wave', 'waterfall', 'rain'],
            'buildings': ['building', 'house', 'skyscraper', 'cottage', 'castle', 'apartment', 'hut', 'structure', 'architecture'],
            # Add more categories as needed
        }

        # Blurriness and low-detail thresholds
        self.blur_threshold = blur_threshold
        self.low_detail_threshold = low_detail_threshold

        # Initialize NLP components
        try:
            self.nlp = spacy.load('en_core_web_sm')
        except ImportError:
            raise ImportError("SpaCy is not installed. Please install it using 'pip install spacy'.")
        except OSError:
            raise OSError("SpaCy model 'en_core_web_sm' is not available. Please install it using 'python -m spacy download en_core_web_sm'.")

    def calc_rectangle(self, xi, yi):
        """Calculate the coordinates for a tile based on its index."""
        x1 = self.tile_positions_x[xi]
        y1 = self.tile_positions_y[yi]

        # For the last column, adjust x2 to not exceed image width
        if xi == self.cols - 1:
            x2 = self.width
        else:
            x2 = x1 + self.tile_size

        # For the last row, adjust y2 to not exceed image height
        if yi == self.rows - 1:
            y2 = self.height
        else:
            y2 = y1 + self.tile_size

        return x1, y1, x2, y2

    def analyze_full_image(self):
        if self.prompt_generator is None:
            self.entire_image_prompt = self.upscaler.p.prompt  # Use main img2img prompt
            return
        print("Generating prompt for the entire image...")
        self.entire_image_prompt = self.prompt_generator.generate_prompt(
            self.upscaler.image, max_words=self.master_prompt_max_words, prefix=self.base_prompt, suffix=self.clip_prompt_suffix
        )

        # Remove words and categories if specified
        self.entire_image_prompt = self.remove_words_from_prompt(self.entire_image_prompt)

        print(f"Entire image prompt: {self.entire_image_prompt}")

    def generate_tile_prompt(self, tile_image):
        if self.prompt_generator is None:
            tile_prompt = f"{self.upscaler.p.prompt}"
        else:
            tile_prompt = self.prompt_generator.generate_prompt(
                tile_image,
                max_words=self.max_clip_words,
                prefix=self.base_prompt,
                suffix=self.clip_prompt_suffix
            )

        # Remove words and categories if specified
        tile_prompt = self.remove_words_from_prompt(tile_prompt)

        tile_prompt = self.clean_and_deduplicate_prompt(tile_prompt)

        return tile_prompt

    def clean_and_deduplicate_prompt(self, prompt):
        """Remove duplicates and unwanted words from the prompt."""
        words = prompt.split()
        unwanted_words = {"he", "him", "his", "a", "the", "and", "in", "of", "with", "on", "at", "from"}
        words = [word for word in words if word.lower() not in unwanted_words]

        # Deduplicate words while preserving order
        seen = set()
        deduplicated_words = []
        for word in words:
            if word.lower() not in seen:
                deduplicated_words.append(word)
                seen.add(word.lower())

        return ' '.join(deduplicated_words)

    def remove_words_from_prompt(self, prompt):
        """Remove specified words and categories from the prompt."""
        words_to_remove = set()

        # Process categories to remove
        for category in self.categories_to_remove:
            category = category.strip().lower()
            if category in self.category_words:
                words_to_remove.update(self.category_words[category])
            else:
                print(f"Category '{category}' not found in predefined categories.")

        # Process individual words to remove
        words_to_remove.update([word.strip().lower() for word in self.words_to_remove])

        if not words_to_remove:
            return prompt

        # Remove words from the prompt
        pattern = re.compile(r'\b(' + '|'.join(map(re.escape, words_to_remove)) + r')\b', re.IGNORECASE)
        prompt = pattern.sub('', prompt)
        prompt = ' '.join(prompt.split())  # Remove extra spaces
        return prompt

    def pre_analyze(self, image):
        """Analyze the full image and generate prompts for each tile."""
        self.analyze_full_image()
        tile_prompts = []

        for yi in range(self.rows):
            row_prompts = []
            for xi in range(self.cols):
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)
                x1 = max(0, min(x1, self.width))
                y1 = max(0, min(y1, self.height))
                x2 = max(0, min(x2, self.width))
                y2 = max(0, min(y2, self.height))
                if x2 <= x1 or y2 <= y1:
                    continue

                tile_image = image.crop((x1, y1, x2, y2))
                tile_prompt = self.generate_tile_prompt(tile_image)
                row_prompts.append(tile_prompt)
            tile_prompts.append(row_prompts)

        return tile_prompts

    def start(self, p, image, tile_prompts):
        self.width, self.height = image.size
        feather_amount = self.feather_amount

        # Create a new image to store the final result
        final_image = Image.new("RGB", (self.width, self.height))

        # Iterate over each tile and process it
        for yi in range(self.rows):
            for xi in range(self.cols):
                # Calculate tile position
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)

                # Ensure tile boundaries are valid
                x1 = max(0, min(x1, self.width))
                y1 = max(0, min(y1, self.height))
                x2 = max(0, min(x2, self.width))
                y2 = max(0, min(y2, self.height))
                if x2 <= x1 or y2 <= y1:
                    continue

                # Crop the tile from the image
                tile_image = image.crop((x1, y1, x2, y2))

                # Get the prompt for this tile
                tile_prompt = tile_prompts[yi][xi]
                print(f"Tile ({xi}, {yi}) prompt: {tile_prompt}")

                # Prepare the processing object for this tile
                tile_p = self.prepare_tile_processing(p, tile_image, tile_prompt)

                # Process the tile
                processed = processing.process_images(tile_p)

                # Get the processed tile
                processed_tile = processed.images[0]

                # Resize the processed tile back to the original tile size
                processed_tile = processed_tile.resize(tile_image.size, Image.LANCZOS)

                # Save the tile if the option is enabled
                if self.save_tiles:
                    tile_filename = f"tile_{yi}_{xi}.png"
                    tile_filepath = os.path.join(p.outpath_samples, tile_filename)
                    processed_tile.save(tile_filepath)
                    print(f"Saved tile: {tile_filepath}")

                # Create a feathered mask
                if feather_amount > 0:
                    mask = self.create_feathered_mask((x1, y1, x2, y2), (self.width, self.height), feather_amount)
                else:
                    mask = Image.new("L", tile_image.size, 255)  # No feathering

                # Ensure mask size matches tile size
                if mask.size != tile_image.size:
                    mask = mask.resize(tile_image.size, Image.LANCZOS)

                # Paste the processed tile back into the final image with blending
                final_image.paste(processed_tile, (x1, y1), mask)

                # Clear CUDA cache to free up memory
                if self.device.type == 'cuda':
                    torch.cuda.empty_cache()

                # Update progress
                state.job_no += 1
                state.job_count = self.upscaler.job_count
                if state.interrupted:
                    break

            if state.interrupted:
                break

        return final_image

    def create_feathered_mask(self, tile_position, image_size, feather_radius):
        tile_width = tile_position[2] - tile_position[0]
        tile_height = tile_position[3] - tile_position[1]
        mask = Image.new("L", (tile_width, tile_height), 255)

        x1, y1, x2, y2 = tile_position
        image_width, image_height = image_size

        # Determine which edges need feathering
        needs_feathering = False
        left = 0
        top = 0
        right = tile_width
        bottom = tile_height

        if x1 > 0:
            left = feather_radius
            needs_feathering = True
        if y1 > 0:
            top = feather_radius
            needs_feathering = True
        if x2 < image_width:
            right = tile_width - feather_radius
            needs_feathering = True
        if y2 < image_height:
            bottom = tile_height - feather_radius
            needs_feathering = True

        if needs_feathering:
            alpha = Image.new("L", (tile_width, tile_height), 0)
            draw = ImageDraw.Draw(alpha)
            draw.rectangle((left, top, right, bottom), fill=255)
            mask = alpha.filter(ImageFilter.GaussianBlur(radius=feather_radius))
        else:
            # The tile is at the image boundary on all sides, no feathering needed
            mask = Image.new("L", (tile_width, tile_height), 255)

        return mask

    def prepare_tile_processing(self, p, tile_image, tile_prompt):
        # Create a shallow copy of the processing object
        tile_p = copy.copy(p)

        # Ensure the width and height are multiples of 64 and do not exceed tile_image size
        tile_p.width = min(tile_image.width, max(64, (tile_image.width // 64) * 64))
        tile_p.height = min(tile_image.height, max(64, (tile_image.height // 64) * 64))

        # Resize the tile_image to match the dimensions required by the model
        resized_tile_image = tile_image.resize((tile_p.width, tile_p.height), Image.LANCZOS)
        tile_p.init_images = [resized_tile_image]
        tile_p.prompt = tile_prompt
        tile_p.batch_size = 1
        tile_p.n_iter = 1
        tile_p.do_not_save_samples = not self.save_tiles
        tile_p.do_not_save_grid = True

        # Fix the seed for reproducibility
        fix_seed(tile_p)

        return tile_p

# USDUpscaler Class
class USDUpscaler:
    model_cache = {}

    def __init__(
        self,
        p,
        image,
        upscaler_index: int,
        save_redraw,
        tile_size,
        base_prompt,
        overlap_percentage,
        clip_prompt_suffix,
        save_prompts_to_file,
        uploaded_prompt_file,
        prompt_method: str,
        global_seed: int,
        master_prompt_max_words: int,
        tile_prompt_max_words: int,
        feather_amount: int,
        blur_threshold: int,
        low_detail_threshold: int,
        save_tiles: bool,  # Parameter for saving tiles
        words_to_remove: str,  # Words to remove from tile prompts
        categories_to_remove: str,  # Categories to remove from tile prompts
        auto_adjust_tile_size: bool  # Auto adjust tile size option
    ) -> None:
        self.p: StableDiffusionProcessing = p
        self.image: Image = image
        # Calculate scale factor based on desired output size
        init_img_width, init_img_height = self.image.size
        desired_width, desired_height = p.width, p.height
        scale_factor_w = desired_width / init_img_width
        scale_factor_h = desired_height / init_img_height

        if abs(scale_factor_w - scale_factor_h) > 0.01:
            print("Warning: Aspect ratio has changed. Proceeding with width scale factor.")

        self.scale_factor = scale_factor_w

        self.upscaler_data = shared.sd_upscalers[upscaler_index]
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")  # Set the device

        # Parse words to remove into a list
        words_to_remove_list = [word.strip().lower() for word in words_to_remove.split(',')] if words_to_remove else []

        # Parse categories to remove into a list
        categories_to_remove_list = [category.strip().lower() for category in categories_to_remove.split(',')] if categories_to_remove else []

        # Initialize the selected prompt generator
        self.prompt_method = prompt_method
        self.prompt_generator = self.initialize_prompt_generator(prompt_method)

        # Initialize the redraw object with the device and thresholds
        self.redraw = USDURedraw(
            usdupscaler=self,
            device=self.device,
            feather_amount=feather_amount,
            save_tiles=save_tiles,
            master_prompt_max_words=master_prompt_max_words,
            tile_prompt_max_words=tile_prompt_max_words,
            blur_threshold=blur_threshold,
            low_detail_threshold=low_detail_threshold,
            prompt_generator=self.prompt_generator,
            words_to_remove=words_to_remove_list,
            categories_to_remove=categories_to_remove_list
        )

        self.redraw.save = save_redraw
        self.redraw.base_prompt = base_prompt
        self.redraw.clip_prompt_suffix = clip_prompt_suffix
        self.initial_info = None
        self.save_prompts_to_file = save_prompts_to_file
        self.uploaded_prompt_file = uploaded_prompt_file
        self.global_seed = global_seed  # Use the parent img2img seed for all tiles

        # Upscale the image using the selected upscaler
        self.upscale_image()

        # Adjust tile size and positions
        self.adjust_tile_size_and_positions(auto_adjust_tile_size, tile_size, overlap_percentage)

    def initialize_prompt_generator(self, method_name: str):
        if method_name == "NONE":
            return None
        if method_name in USDUpscaler.model_cache:
            return USDUpscaler.model_cache[method_name]

        try:
            if method_name == "BLIP":
                generator = BLIPPromptGenerator(self.device)
            elif method_name == "CLIP":
                generator = CLIPPromptGenerator(self.device)
            else:
                raise ValueError(f"Unsupported prompt method: {method_name}")

            USDUpscaler.model_cache[method_name] = generator
            return generator
        except Exception as e:
            print(f"Error initializing prompt generator '{method_name}': {e}")
            raise e

    def upscale_image(self):
        # Upscale the image using the selected upscaler
        if self.upscaler_data:
            print(f"Upscaling image using {self.upscaler_data.name}...")
            self.image = self.upscaler_data.scaler.upscale(self.image, self.scale_factor, self.upscaler_data.data_path)
            # Update p.width and p.height to match the upscaled image
            self.p.width, self.p.height = self.image.size
            print(f"Image upscaled to {self.p.width}x{self.p.height}")
        else:
            print("No upscaler selected. Skipping upscaling step.")

    def adjust_tile_size_and_positions(self, auto_adjust_tile_size, tile_size, overlap_percentage):
        width, height = self.image.size

        # Auto adjust tile size if enabled
        if auto_adjust_tile_size:
            # Calculate the overlap in pixels (as percentage of the tile size)
            overlap = int(overlap_percentage / 100 * tile_size)

            # Calculate the number of tiles needed along each dimension
            self.cols = max(1, math.ceil((width + overlap) / (tile_size - overlap)))
            self.rows = max(1, math.ceil((height + overlap) / (tile_size - overlap)))

            # Recalculate tile size to fit the image exactly
            effective_tile_size_x = (width + (self.cols - 1) * overlap) // self.cols
            effective_tile_size_y = (height + (self.rows - 1) * overlap) // self.rows
            self.redraw.tile_size = min(effective_tile_size_x, effective_tile_size_y)

            print(f"Auto-adjusted tile size to {self.redraw.tile_size}")
        else:
            self.redraw.tile_size = tile_size
            print(f"Using specified tile size {self.redraw.tile_size}")

        # Calculate overlap in pixels
        self.overlap = int(overlap_percentage / 100 * self.redraw.tile_size)
        self.redraw.overlap_percentage = overlap_percentage

        # Recalculate the number of tiles to cover the image without partial tiles
        self.cols = max(1, math.ceil((width + self.overlap) / (self.redraw.tile_size - self.overlap)))
        self.rows = max(1, math.ceil((height + self.overlap) / (self.redraw.tile_size - self.overlap)))

        # Generate tile positions
        self.redraw.tile_positions_x = [min(i * (self.redraw.tile_size - self.overlap), width - self.redraw.tile_size) for i in range(self.cols)]
        self.redraw.tile_positions_y = [min(i * (self.redraw.tile_size - self.overlap), height - self.redraw.tile_size) for i in range(self.rows)]
        self.redraw.cols = self.cols
        self.redraw.rows = self.rows
        self.redraw.width = width
        self.redraw.height = height

        print(f"Adjusted tile size to {self.redraw.tile_size} to fit image dimensions without partial tiles.")
        print(f"Number of tiles: {self.cols} cols x {self.rows} rows")

    def pre_analyze_tiles(self):
        # Pre-analyze each tile and generate prompts
        if self.uploaded_prompt_file:
            # Load prompts from Excel file
            self.tile_prompts = self.load_prompts_from_excel(self.uploaded_prompt_file)
        else:
            self.tile_prompts = self.redraw.pre_analyze(self.image)
            if self.save_prompts_to_file and self.prompt_generator is not None:
                self.save_prompts_to_excel(self.tile_prompts)

    def calc_jobs_count(self):
        # Calculate the number of jobs (tiles) to process
        self.job_count = self.redraw.rows * self.redraw.cols
        state.job_count = self.job_count

    def add_extra_info(self):
        # Add any extra info to initial_info
        self.initial_info = "Upscaling complete."

    def process(self):
        state.begin()
        self.calc_jobs_count()
        self.result_images = []

        self.pre_analyze_tiles()

        # Start processing with the prompts
        self.image = self.redraw.start(self.p, self.image, self.tile_prompts)
        self.initial_info = self.redraw.initial_info

        self.result_images.append(self.image)
        if self.redraw.save:
            print("Saving the final upscaled image...")
            self.save_image()

        # Cleanup
        del self.image
        del self.redraw

        state.end()

    def save_image(self):
        # Implement image saving logic
        method_suffix = f"_{self.prompt_method}"
        images.save_image(
            self.image,
            self.p.outpath_samples,
            "",  # filename
            seed=self.global_seed,
            prompt=self.p.prompt,
            extension=opts.samples_format,
            info=self.initial_info,
            p=self.p,
            suffix=method_suffix  # Add method name as suffix
        )

    def save_prompts_to_excel(self, tile_prompts):
        # Save the prompts to an Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Tile Prompts"

        # Write prompts to the worksheet
        for yi, row_prompts in enumerate(tile_prompts):
            for xi, prompt in enumerate(row_prompts):
                cell = ws.cell(row=yi+1, column=xi+1)
                cell.value = prompt
                cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value) if column_cells else 0
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        # Save the workbook
        excel_file_base = os.path.join(self.p.outpath_samples, f"tile_prompts_{self.prompt_method}")
        excel_file_path = excel_file_base + ".xlsx"
        suffix = 1
        while True:
            try:
                wb.save(excel_file_path)
                print(f"Tile prompts saved to {excel_file_path}")
                break
            except PermissionError:
                # File is open or cannot be written over, try a new filename
                excel_file_path = excel_file_base + f"_{suffix}.xlsx"
                suffix += 1

    def load_prompts_from_excel(self, uploaded_file):
        # Load prompts from an uploaded Excel file
        wb = load_workbook(filename=uploaded_file.name)
        ws = wb.active

        tile_prompts = []
        for row in ws.iter_rows(values_only=True):
            row_prompts = [cell if cell is not None else "" for cell in row]
            tile_prompts.append(row_prompts)

        # Adjust rows and columns if necessary
        expected_rows = self.redraw.rows
        expected_cols = self.redraw.cols
        actual_rows = len(tile_prompts)
        actual_cols = max(len(row) for row in tile_prompts) if tile_prompts else 0

        if actual_rows != expected_rows or actual_cols != expected_cols:
            raise ValueError(f"Excel file dimensions ({actual_rows}x{actual_cols}) do not match expected tile grid ({expected_rows}x{expected_cols}).")

        return tile_prompts

# Script Class for Gradio UI
class Script(scripts.Script):
    def title(self):
        return "Regional Prompt Upscaler - Free Version"

    def show(self, is_img2img):
        return is_img2img

    def ui(self, is_img2img):
        # Define elem_id_prefix to avoid NameError
        elem_id_prefix = "regional_upscaler"

        # Configurable text for easy editing
        version_text = "Regional Prompt Upscaler - Free Version"
        website_text = "hallett-ai.com"
        free_version_message = "Check out my Pro version for more features."

        # Displaying the configurable info in the UI
        version_info = gr.HTML(f"<p style=\"font-weight:bold;\">{version_text}</p>")
        website_info = gr.HTML(f"<p>{website_text}</p>")
        free_message = gr.HTML(f"<p>{free_version_message}</p>")

        # Upscaler settings section
        info = gr.HTML("<p>Will upscale the image using the img2img scale setting.</p>")

        # Upscaler model selection
        with gr.Row():
            upscaler_index = gr.Radio(
                label='Upscaler Model',
                elem_id=f"{elem_id_prefix}_upscaler_index",
                choices=[x.name for x in shared.sd_upscalers],
                value=shared.sd_upscalers[0].name,
                type="index"
            )

        # Tile settings section
        with gr.Row():
            tile_size = gr.Slider(
                elem_id=f"{elem_id_prefix}_tile_size",
                minimum=256,
                maximum=2048,
                step=64,
                label='Tile size',
                value=1024
            )
            feather_amount = gr.Slider(
                elem_id=f"{elem_id_prefix}_feather_amount",
                label='Feather Amount',
                minimum=0,
                maximum=100,
                step=1,
                value=50
            )
            overlap_percentage = gr.Slider(
                label='Overlap Percentage',
                minimum=0,
                maximum=50,
                step=1,
                value=20
            )
        with gr.Row():
            auto_adjust_tile_size = gr.Checkbox(
                label='Auto Adjust Tile Size',
                value=True
            )

        # Prompt generation settings
        prompt_methods = ["BLIP", "CLIP", "NONE"]

        gr.HTML("<p>Prompt Generation Settings:</p>")
        with gr.Row():
            base_prompt = gr.Textbox(
                label="Prefix (Base Prompt)",
                placeholder="Enter a base prompt...",
                lines=1
            )
            clip_prompt_suffix = gr.Textbox(
                label="Suffix",
                placeholder="Enter a suffix for prompts...",
                lines=1
            )
        with gr.Row():
            prompt_method = gr.Dropdown(
                label="Prompt Generation Method",
                elem_id=f"{elem_id_prefix}_prompt_method",
                choices=prompt_methods,
                value="BLIP"
            )
            master_prompt_max_words = gr.Slider(
                label='Master Prompt Max Words',
                minimum=5,
                maximum=50,
                step=1,
                value=20
            )
            tile_prompt_max_words = gr.Slider(
                label='Tile Prompt Max Words',
                minimum=1,
                maximum=20,
                step=1,
                value=15
            )
        with gr.Row():
            words_to_remove = gr.Textbox(
                label="Words to Remove from Tile Prompts",
                placeholder="Enter words to remove, separated by commas...",
                lines=1
            )
            categories_to_remove = gr.Textbox(
                label="Categories to Remove from Tile Prompts",
                placeholder="Enter categories to remove, separated by commas...",
                lines=1
            )

        # File save settings
        gr.HTML("<p>Save options:</p>")
        with gr.Row():
            save_prompts_to_file = gr.Checkbox(
                label='Save Prompts to Excel File',
                value=True
            )
            upload_prompt_file = gr.File(
                label='Upload Prompt Excel File',
                file_types=['.xlsx']
            )

        # Threshold settings section
        gr.HTML("<p>Threshold Settings:</p>")
        with gr.Row():
            blur_threshold = gr.Slider(
                label='Blurriness Detection Level',
                minimum=0,
                maximum=10,
                step=1,
                value=1
            )
            low_detail_threshold = gr.Slider(
                label='Low Detail Detection Level',
                minimum=0,
                maximum=10,
                step=1,
                value=1
            )

        # Save options for images and tiles
        with gr.Row():
            save_upscaled_image = gr.Checkbox(
                label="Save Upscaled Image",
                elem_id=f"{elem_id_prefix}_save_upscaled_image",
                value=True
            )
            save_tiles = gr.Checkbox(
                label="Save Each Tile",
                value=False
            )

        # Return only the components needed for the run method
        return [
            tile_size,
            feather_amount,
            upscaler_index, 
            base_prompt, 
            clip_prompt_suffix, 
            prompt_method, 
            master_prompt_max_words, 
            tile_prompt_max_words, 
            save_prompts_to_file, 
            upload_prompt_file, 
            overlap_percentage, 
            blur_threshold, 
            low_detail_threshold, 
            save_upscaled_image, 
            save_tiles, 
            words_to_remove, 
            categories_to_remove, 
            auto_adjust_tile_size
        ]

    def run(self, p, *args):
        (
            tile_size,
            feather_amount,
            upscaler_index, 
            base_prompt, 
            clip_prompt_suffix, 
            prompt_method, 
            master_prompt_max_words, 
            tile_prompt_max_words, 
            save_prompts_to_file, 
            upload_prompt_file,  # Changed from 'uploaded_prompt_file' to match 'upload_prompt_file'
            overlap_percentage, 
            blur_threshold, 
            low_detail_threshold, 
            save_upscaled_image, 
            save_tiles, 
            words_to_remove, 
            categories_to_remove, 
            auto_adjust_tile_size
        ) = args

        # Assign the fixed seed immediately before any processing
        fix_seed(p)
        global_seed = p.seed

        # Initialize image
        init_img = p.init_images[0]
        if init_img is None:
            return Processed(p, [], p.seed, "Empty image")
        init_img = images.flatten(init_img, opts.img2img_background_color)

        # Upscaling process
        upscaler = USDUpscaler(
            p=p,
            image=init_img,
            upscaler_index=upscaler_index,
            save_redraw=save_upscaled_image,
            tile_size=tile_size,
            tile_prompt_max_words=tile_prompt_max_words,  # Changed here
            base_prompt=base_prompt,
            overlap_percentage=overlap_percentage,
            clip_prompt_suffix=clip_prompt_suffix,
            save_prompts_to_file=save_prompts_to_file,
            uploaded_prompt_file=upload_prompt_file,  # Ensure consistency if variable names differ
            prompt_method=prompt_method,
            global_seed=global_seed,
            master_prompt_max_words=master_prompt_max_words,
            feather_amount=feather_amount,
            blur_threshold=blur_threshold,
            low_detail_threshold=low_detail_threshold,
            save_tiles=save_tiles,
            words_to_remove=words_to_remove,
            categories_to_remove=categories_to_remove,
            auto_adjust_tile_size=auto_adjust_tile_size
        )
        
        upscaler.add_extra_info()
        upscaler.process()
        result_images = upscaler.result_images

        return Processed(p, result_images, p.seed, upscaler.initial_info or "")
