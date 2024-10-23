# Regional Prompt Upscaler - Free Version v1.10

import subprocess
import sys
import importlib
import hashlib
import io

# Utility functions to install missing packages
def install(package):
    try:
        print(f"Installing package: {package}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    except Exception as e:
        print(f"Failed to install {package}: {e}")

def ensure_package_installed(package, module_name=None):
    if module_name is None:
        module_name = package
    try:
        __import__(module_name)
    except ImportError:
        print(f"{package} is not installed. Attempting to install...")
        install(package)

# Ensure required packages are installed
def install_dependencies():
    ensure_package_installed('torch')
    ensure_package_installed('transformers')
    ensure_package_installed('openpyxl')  # Ensure openpyxl is installed
    ensure_package_installed('numpy')
    ensure_package_installed('Pillow', module_name='PIL')
    ensure_package_installed('gradio')

# Install dependencies before imports
install_dependencies()

# Now, import all the necessary modules
import torch
import math
import gradio as gr
import numpy as np
from PIL import Image, ImageDraw, ImageFilter
from modules import processing, shared, images, scripts
from modules.processing import Processed, fix_seed
from modules.shared import opts, state
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import re
import copy
import gc

from transformers import BlipProcessor, BlipForConditionalGeneration

# Initialize device
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# Base class for Prompt Generators
class PromptGeneratorBase:
    def __init__(self, device: torch.device):
        self.device = device

    def generate_prompt(self, image: Image.Image, prompt_mode: str = "Complex", max_words: int = None, prefix: str = "", suffix: str = "") -> str:
        raise NotImplementedError("generate_prompt method must be implemented by subclasses.")

    def supports_save_load(self) -> bool:
        """Indicates whether this generator supports prompt saving and loading."""

# BLIP Prompt Generator
class BLIPPromptGenerator(PromptGeneratorBase):
    def __init__(self, device: torch.device, model_name: str = "Salesforce/blip-image-captioning-base"):
        super().__init__(device)
        self.processor = BlipProcessor.from_pretrained(model_name)
        self.model = BlipForConditionalGeneration.from_pretrained(
            model_name,
            torch_dtype=torch.float16 if device.type == 'cuda' else torch.float32
        ).to(self.device)
        self.model.eval()

    def generate_prompt(self, image: Image.Image, max_words: int = None, prefix: str = "", suffix: str = "") -> str:
        try:
            # Use an empty string to let BLIP generate captions without a predefined context
            prompt_text = ""

            # Preprocess the image and text
            inputs = self.processor(images=image, text=prompt_text, return_tensors="pt").to(self.device)

            # Generate the caption
            with torch.no_grad():
                outputs = self.model.generate(**inputs, max_length=100, num_beams=5, no_repeat_ngram_size=2, early_stopping=True)

            # Decode the generated caption
            caption = self.processor.decode(outputs[0], skip_special_tokens=True)

            # Optionally trim based on max_words
            if max_words:
                caption = self._trim_caption(caption, max_words=max_words)

            # Construct the final prompt (without prefix like "Describe this image in detail")
            prompt = f"{prefix.strip()} {caption} {suffix.strip()}".strip()
            return prompt

        except Exception as e:
            print(f"Error generating prompt with BLIP: {e}")
            return "an image"

# USDURedraw Class
class USDURedraw:
    def __init__(
        self,
        usdupscaler,
        device,
        feather_amount=50,
        save_tiles=False,
        exclude_humans_animals=False,
        prompt_generator=None,
        words_to_remove=None,
        categories_to_remove=None,
        auto_adjust_tile_size=False
    ):
        self.upscaler = usdupscaler
        self.device = device
        self.save_tiles = save_tiles
        self.tile_size = 1024
        self.enabled = True
        self.base_prompt = ""
        self.initial_info = None
        self.overlap_percentage = 0
        self.clip_prompt_suffix = ""
        self.entire_image_prompt = ""
        self.feather_amount = feather_amount
        self.prompt_generator = prompt_generator
        self.words_to_remove = words_to_remove or []
        self.categories_to_remove = categories_to_remove or []
        self.auto_adjust_tile_size = auto_adjust_tile_size

        # Terms related to humans that can be excluded
        self.human_terms = {
            "human", "person", "man", "woman", "girl", "boy", "child", "baby",
            "hand", "face", "people", "male", "female", "lady", "gentleman", "kid",
            "body", "head", "arm", "leg", "finger", "foot", "hair", "eye", "mouth",
            "nose", "ear", "neck", "shoulder", "portrait", "skin", "smile"
        }
        self.exclude_humans_animals = exclude_humans_animals

        # Predefined category words for removal
        self.category_words = {
            'animals': ['animal', 'dog', 'cat', 'bird', 'lion', 'tiger', 'elephant', 'horse', 'fish', 'bear', 'wolf', 'rabbit', 'fox', 'deer', 'cow', 'sheep', 'goat', 'monkey', 'kangaroo', 'panda'],
            'sky': ['sky', 'cloud', 'sun', 'moon', 'star', 'sunset', 'sunrise', 'aurora', 'galaxy'],
            'water': ['water', 'lake', 'river', 'ocean', 'sea', 'wave', 'waterfall', 'beach', 'shore', 'coast'],
            'buildings': ['building', 'house', 'skyscraper', 'castle', 'apartment', 'temple', 'church', 'palace', 'factory', 'bridge', 'tower'],
            'vehicles': ['car', 'truck', 'bus', 'train', 'bicycle', 'motorcycle', 'airplane', 'boat', 'submarine'],
            'plants': ['tree', 'flower', 'grass', 'forest', 'garden', 'leaf', 'vine'],
            'food': ['fruit', 'vegetable', 'meal', 'dish', 'dessert', 'snack', 'breakfast', 'lunch', 'dinner']
        }

    def calc_rectangle(self, xi, yi):
        """Calculate the coordinates for a tile based on its index."""
        x1 = self.tile_positions_x[xi]
        y1 = self.tile_positions_y[yi]

        # For the last column, adjust x2 to not exceed image width
        if xi == self.cols - 1:
            x2 = self.width
        else:
            x2 = x1 + self.tile_size

        # For the last row, adjust y2 to not exceed image height
        if yi == self.rows - 1:
            y2 = self.height
        else:
            y2 = y1 + self.tile_size

        print(f"Tile ({xi}, {yi}) position: ({x1}, {y1}, {x2}, {y2})")

        return x1, y1, x2, y2

    def analyze_full_image(self):
        """Generate a prompt for the entire image using BLIP."""
        if self.prompt_generator is None:
            self.entire_image_prompt = self.upscaler.p.prompt or ""
        else:
            print("Generating prompt for the entire image...")
            self.entire_image_prompt = self.prompt_generator.generate_prompt(
                self.upscaler.image, prefix=self.base_prompt, suffix=self.clip_prompt_suffix
            )

        if self.exclude_humans_animals:
            self.entire_image_prompt = self.remove_human_terms(self.entire_image_prompt)

        print(f"Entire image prompt: {self.entire_image_prompt}")

    def generate_tile_prompt(self, tile_image):
        """Generate a prompt for an individual tile."""
        if self.prompt_generator is None:
            tile_prompt = self.upscaler.p.prompt
        else:
            print("Generating prompt for tile...")
            tile_prompt = self.prompt_generator.generate_prompt(
                tile_image, prefix=self.base_prompt, suffix=self.clip_prompt_suffix
            )

        if self.exclude_humans_animals:
            tile_prompt = self.remove_human_terms(tile_prompt)

        tile_prompt = self.remove_words_from_prompt(tile_prompt)
        tile_prompt = self.remove_border_influence(tile_prompt)

        return tile_prompt

    def remove_human_terms(self, prompt):
        # Use regex to remove human-related terms
        pattern = re.compile(r'\b(' + '|'.join(map(re.escape, self.human_terms)) + r')\b', re.IGNORECASE)
        prompt = pattern.sub('', prompt)
        return ' '.join(prompt.split())  # Remove extra spaces

    def remove_words_from_prompt(self, prompt):
        words_to_remove = set()

        # Process categories to remove
        for category in self.categories_to_remove:
            category = category.strip().lower()
            if category in self.category_words:
                words_to_remove.update(self.category_words[category])

        # Process individual words to remove
        words_to_remove.update([word.strip().lower() for word in self.words_to_remove])

        if not words_to_remove:
            return prompt

        # Match singular and plural forms automatically (e.g., "animal" and "animals")
        pattern = re.compile(r'\b(' + '|'.join([f"{re.escape(word)}s?" for word in words_to_remove]) + r')\b', re.IGNORECASE)
        prompt = pattern.sub('', prompt)
        return ' '.join(prompt.split())  # Remove extra spaces

    def remove_border_influence(self, prompt):
        """Remove incorrect descriptions caused by feathering and overlap."""
        return re.sub(r'\bwhite border\b', '', prompt, flags=re.IGNORECASE).strip()

    def pre_analyze(self, image):
        """Analyze the full image and generate prompts for each tile."""
        self.analyze_full_image()
        tile_prompts = []

        for yi in range(self.rows):
            row_prompts = []
            for xi in range(self.cols):
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)
                x1 = max(0, min(x1, self.width))
                y1 = max(0, min(y1, self.height))
                x2 = max(0, min(x2, self.width))
                y2 = max(0, min(y2, self.height))
                if x2 <= x1 or y2 <= y1:
                    continue

                tile_image = image.crop((x1, y1, x2, y2))
                tile_prompt = self.generate_tile_prompt(tile_image)
                row_prompts.append(tile_prompt)
            tile_prompts.append(row_prompts)

        return tile_prompts

    def start(self, p, image, tile_prompts):
        self.width, self.height = image.size
        feather_amount = self.feather_amount

        # Create a new image to store the final result
        final_image = Image.new("RGB", (self.width, self.height))

        # Iterate over each tile and process it
        for yi in range(self.rows):
            for xi in range(self.cols):
                # Calculate tile position
                x1, y1, x2, y2 = self.calc_rectangle(xi, yi)

                # Ensure tile boundaries are valid
                x1 = max(0, min(x1, self.width))
                y1 = max(0, min(y1, self.height))
                x2 = max(0, min(x2, self.width))
                y2 = max(0, min(y2, self.height))
                if x2 <= x1 or y2 <= y1:
                    continue

                # Crop the tile from the image
                tile_image = image.crop((x1, y1, x2, y2))

                # Get the prompt for this tile
                tile_prompt = tile_prompts[yi][xi]
                print(f"Tile ({xi}, {yi}) prompt: {tile_prompt}")

                # Prepare the processing object for this tile
                tile_p = self.prepare_tile_processing(p, tile_image, tile_prompt)

                # Process the tile
                processed = processing.process_images(tile_p)

                # Get the processed tile
                processed_tile = processed.images[0]

                # Resize the processed tile back to the original tile size
                processed_tile = processed_tile.resize(tile_image.size, Image.LANCZOS)

                # Save the tile if the option is enabled
                if self.save_tiles:
                    tile_filename = f"tile_{yi}_{xi}.png"
                    tile_filepath = os.path.join(p.outpath_samples, tile_filename)
                    processed_tile.save(tile_filepath)
                    print(f"Saved tile: {tile_filepath}")

                # Create a feathered mask
                if feather_amount > 0:
                    mask = self.create_feathered_mask((x1, y1, x2, y2), (self.width, self.height), feather_amount)
                else:
                    mask = Image.new("L", tile_image.size, 255)  # No feathering

                # Ensure mask size matches tile size
                if mask.size != tile_image.size:
                    mask = mask.resize(tile_image.size, Image.LANCZOS)

                # Paste the processed tile back into the final image with blending
                final_image.paste(processed_tile, (x1, y1), mask)

                # Clear CUDA cache to free up memory
                if self.device.type == 'cuda':
                    torch.cuda.empty_cache()

                # Delete variables to free memory
                del tile_image, processed, processed_tile, mask, tile_p

                # Run garbage collection
                gc.collect()

                # Update progress
                state.job_no += 1
                state.job_count = self.upscaler.job_count
                if state.interrupted:
                    break

            if state.interrupted:
                break

        return final_image

    def create_feathered_mask(self, tile_position, image_size, feather_radius):
        """Create a feathered mask to blend tile borders."""
        tile_width = tile_position[2] - tile_position[0]
        tile_height = tile_position[3] - tile_position[1]
        mask = Image.new("L", (tile_width, tile_height), 255)

        x1, y1, x2, y2 = tile_position
        image_width, image_height = image_size

        # Determine which edges need feathering
        needs_feathering = False
        left = 0
        top = 0
        right = tile_width
        bottom = tile_height

        if x1 > 0:
            left = feather_radius
            needs_feathering = True
        if y1 > 0:
            top = feather_radius
            needs_feathering = True
        if x2 < image_width:
            right = tile_width - feather_radius
            needs_feathering = True
        if y2 < image_height:
            bottom = tile_height - feather_radius
            needs_feathering = True

        if needs_feathering:
            alpha = Image.new("L", (tile_width, tile_height), 0)
            draw = ImageDraw.Draw(alpha)
            draw.rectangle((left, top, right, bottom), fill=255)
            mask = alpha.filter(ImageFilter.GaussianBlur(radius=feather_radius))
        else:
            mask = Image.new("L", (tile_width, tile_height), 255)

        return mask

    def prepare_tile_processing(self, p, tile_image, tile_prompt):
        """Prepare the processing object for each tile."""
        tile_p = copy.copy(p)

        # Ensure the width and height are multiples of 64 and do not exceed tile size
        tile_p.width = min(tile_image.width, max(64, (tile_image.width // 64) * 64))
        tile_p.height = min(tile_image.height, max(64, (tile_image.height // 64) * 64))

        # Resize the tile image to match the dimensions required by the model
        resized_tile_image = tile_image.resize((tile_p.width, tile_p.height), Image.LANCZOS)
        tile_p.init_images = [resized_tile_image]
        tile_p.prompt = tile_prompt
        tile_p.batch_size = 1
        tile_p.n_iter = 1
        tile_p.do_not_save_samples = not self.save_tiles
        tile_p.do_not_save_grid = True

        # Fix the seed for reproducibility
        fix_seed(tile_p)

        return tile_p

class USDUpscaler:
    model_cache = {}

    def __init__(
        self,
        p,
        image,
        upscaler_index: int,
        save_redraw,
        tile_size,
        base_prompt,
        overlap_percentage,
        clip_prompt_suffix,
        prompt_method: str,
        prompt_mode_tile: str,
        save_prompts_to_file,
        uploaded_prompt_file,
        global_seed: int,
        feather_amount: int,
        save_tiles: bool,
        exclude_humans_animals: bool,
        words_to_remove: str,
        categories_to_remove: str,
        auto_adjust_tile_size: bool
    ) -> None:
        self.p = p
        self.image: Image = image
        self.prompt_mode_tile = prompt_mode_tile

        # Calculate scale factor based on desired output size
        init_img_width, init_img_height = self.image.size
        desired_width, desired_height = p.width, p.height
        scale_factor_w = desired_width / init_img_width
        scale_factor_h = desired_height / init_img_height

        if abs(scale_factor_w - scale_factor_h) > 0.01:
            print("Warning: Aspect ratio has changed. Proceeding with width scale factor.")

        self.scale_factor = scale_factor_w

        self.upscaler_data = shared.sd_upscalers[upscaler_index]
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")  # Set the device

        # Parse words to remove into a list
        words_to_remove_list = [word.strip().lower() for word in words_to_remove.split(',')] if words_to_remove else []

        # Parse categories to remove into a list
        categories_to_remove_list = [category.strip().lower() for category in categories_to_remove.split(',')] if categories_to_remove else []

        # Initialize the selected prompt generator
        self.prompt_method = prompt_method
        self.prompt_generator = self.initialize_prompt_generator(prompt_method)

        # Initialize the redraw object with the device and thresholds
        self.redraw = USDURedraw(
            usdupscaler=self,
            device=self.device,
            feather_amount=feather_amount,
            save_tiles=save_tiles,
            exclude_humans_animals=exclude_humans_animals,
            prompt_generator=self.prompt_generator,
            words_to_remove=words_to_remove_list,
            categories_to_remove=categories_to_remove_list,
            auto_adjust_tile_size=auto_adjust_tile_size
        )

        self.redraw.save = save_redraw
        self.redraw.base_prompt = base_prompt
        self.redraw.clip_prompt_suffix = clip_prompt_suffix
        self.initial_info = None
        self.save_prompts_to_file = save_prompts_to_file
        self.uploaded_prompt_file = uploaded_prompt_file
        self.global_seed = global_seed  # Use the parent img2img seed for all tiles

        # Upscale the image using the selected upscaler
        self.upscale_image()

        # Adjust tile size and positions
        self.adjust_tile_size_and_positions(auto_adjust_tile_size, tile_size, overlap_percentage)

    def initialize_prompt_generator(self, method_name: str):
        if method_name == "NONE":
            return None
        if method_name in USDUpscaler.model_cache:
            return USDUpscaler.model_cache[method_name]

        try:
            if method_name == "BLIP":
                generator = BLIPPromptGenerator(self.device)
            else:
                raise ValueError(f"Unsupported prompt method: {method_name}")

            USDUpscaler.model_cache[method_name] = generator
            return generator
        except Exception as e:
            print(f"Error initializing prompt generator '{method_name}': {e}")
            raise e

    def upscale_image(self):
        # Upscale the image using the selected upscaler
        if self.upscaler_data:
            print(f"Upscaling image using {self.upscaler_data.name}...")
            self.image = self.upscaler_data.scaler.upscale(self.image, self.scale_factor, self.upscaler_data.data_path)
            # Update p.width and p.height to match the upscaled image
            self.p.width, self.p.height = self.image.size
            print(f"Image upscaled to {self.p.width}x{self.p.height}")
        else:
            print("No upscaler selected. Skipping upscaling step.")

    def adjust_tile_size_and_positions(self, auto_adjust_tile_size, tile_size, overlap_percentage):
        width, height = self.image.size

        # Auto adjust tile size if enabled
        if auto_adjust_tile_size:
            overlap = int(overlap_percentage / 100 * tile_size)
            self.cols = max(1, math.ceil((width + overlap) / (tile_size - overlap)))
            self.rows = max(1, math.ceil((height + overlap) / (tile_size - overlap)))

            effective_tile_size_x = (width + (self.cols - 1) * overlap) // self.cols
            effective_tile_size_y = (height + (self.rows - 1) * overlap) // self.rows
            self.redraw.tile_size = min(effective_tile_size_x, effective_tile_size_y)

            print(f"Auto-adjusted tile size to {self.redraw.tile_size}")
        else:
            self.redraw.tile_size = tile_size
            print(f"Using specified tile size {self.redraw.tile_size}")

        self.overlap = int(overlap_percentage / 100 * self.redraw.tile_size)
        self.redraw.overlap_percentage = overlap_percentage

        self.cols = max(1, math.ceil((width + self.overlap) / (self.redraw.tile_size - self.overlap)))
        self.rows = max(1, math.ceil((height + self.overlap) / (self.redraw.tile_size - self.overlap)))

        # Correct tile positions to prevent duplication
        self.redraw.tile_positions_x = []
        for i in range(self.cols):
            if i == self.cols - 1:
                pos = width - self.redraw.tile_size
            else:
                pos = i * (self.redraw.tile_size - self.overlap)
            self.redraw.tile_positions_x.append(pos)

        self.redraw.tile_positions_y = []
        for i in range(self.rows):
            if i == self.rows - 1:
                pos = height - self.redraw.tile_size
            else:
                pos = i * (self.redraw.tile_size - self.overlap)
            self.redraw.tile_positions_y.append(pos)

        self.redraw.cols = self.cols
        self.redraw.rows = self.rows
        self.redraw.width = width
        self.redraw.height = height

        print(f"Adjusted tile size to {self.redraw.tile_size} to fit image dimensions without partial tiles.")
        print(f"Number of tiles: {self.cols} cols x {self.rows} rows")


    def pre_analyze_tiles(self):
        # Pre-analyze each tile and generate prompts
        if self.uploaded_prompt_file:
            # Load prompts from Excel file
            self.tile_prompts = self.load_prompts_from_excel(self.uploaded_prompt_file)
        else:
            # Analyze image to generate tile prompts with the appropriate mode (Simple/Complex)
            self.tile_prompts = self.redraw.pre_analyze(self.image)
            if self.save_prompts_to_file and self.prompt_generator is not None:
                self.save_prompts_to_excel(self.tile_prompts)

    def calc_jobs_count(self):
        # Calculate the number of jobs (tiles) to process
        self.job_count = self.redraw.rows * self.redraw.cols
        state.job_count = self.job_count

    def add_extra_info(self):
        # Add any extra info to initial_info
        self.initial_info = "Upscaling complete."

    def process(self):
        state.begin()
        self.calc_jobs_count()
        self.result_images = []

        self.pre_analyze_tiles()

        # Start processing with the prompts
        self.image = self.redraw.start(self.p, self.image, self.tile_prompts)
        self.initial_info = self.redraw.initial_info

        self.result_images.append(self.image)
        if self.redraw.save:
            print("Saving the final upscaled image...")
            self.save_image()

        # Cleanup
        self.finalize_processing()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

        state.end()

    def finalize_processing(self):
        # Delete prompt generator instance
        if self.prompt_generator:
            del self.prompt_generator

        # Clear CUDA cache and run garbage collection
        torch.cuda.empty_cache()
        gc.collect()

    def save_image(self):
        # Implement image saving logic
        method_suffix = f"_{self.prompt_method}"
        images.save_image(
            self.result_images[0],
            self.p.outpath_samples,
            "",  # filename
            seed=self.global_seed,
            prompt=self.p.prompt,
            extension=opts.samples_format,
            info=self.initial_info,
            p=self.p,
            suffix=method_suffix
        )

    def save_prompts_to_excel(self, tile_prompts):
        # Save the prompts to an Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Tile Prompts"

        # Write prompts to the worksheet
        for yi, row_prompts in enumerate(tile_prompts):
            for xi, prompt in enumerate(row_prompts):
                cell = ws.cell(row=yi+1, column=xi+1)
                cell.value = prompt
                cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value) if column_cells else 0
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        # Save the workbook
        excel_file_base = os.path.join(self.p.outpath_samples, f"tile_prompts_{self.prompt_method}")
        excel_file_path = excel_file_base + ".xlsx"
        suffix = 1
        while True:
            try:
                wb.save(excel_file_path)
                print(f"Tile prompts saved to {excel_file_path}")
                break
            except PermissionError:
                excel_file_path = excel_file_base + f"_{suffix}.xlsx"
                suffix += 1

    def load_prompts_from_excel(self, uploaded_file):
        # Load prompts from an uploaded Excel file
        if uploaded_file is None:
            raise ValueError("No Excel file uploaded.")

        try:
            # If `uploaded_file` is a file-like object (in-memory), read it directly
            wb = load_workbook(filename=uploaded_file.name) if hasattr(uploaded_file, 'name') else load_workbook(uploaded_file)
            ws = wb.active
        except Exception as e:
            raise ValueError(f"Error loading Excel file: {e}")

        tile_prompts = []
        for row in ws.iter_rows(values_only=True):
            # Handle empty rows more gracefully
            if row is None or all(cell is None for cell in row):
                continue
            row_prompts = [cell if cell is not None else "" for cell in row]
            tile_prompts.append(row_prompts)

        expected_rows = self.redraw.rows
        expected_cols = self.redraw.cols
        actual_rows = len(tile_prompts)
        actual_cols = max(len(row) for row in tile_prompts) if tile_prompts else 0

        # Debug print to inspect tile grid dimensions
        print(f"Expected (rows, cols): ({expected_rows}, {expected_cols})")
        print(f"Actual (rows, cols): ({actual_rows}, {actual_cols})")

        if actual_rows != expected_rows or actual_cols != expected_cols:
            raise ValueError(f"Excel file dimensions ({actual_rows}x{actual_cols}) do not match expected tile grid ({expected_rows}x{expected_cols}).")

        return tile_prompts

# Script Class for Gradio UI
class Script(scripts.Script):
    def title(self):
        return "Regional Prompt Upscaler - Free Version v1.10"

    def show(self, is_img2img):
        return is_img2img

    def ui(self, is_img2img):
        elem_id_prefix = "regional_upscaler_free"

        version_info = gr.HTML(f"<p style=\"font-weight:bold;\">Regional Prompt Upscaler - Free Version v1.10</p>")
        website_info = gr.HTML(f"<p>hallett-ai.com</p>")
        pro_message = gr.HTML(f"<p>The Pro Version Contains Complex Prompt Generators That Are More Accurate</p>")

        info = gr.HTML("<p>Will upscale the image using the img2img scale setting with advanced options available.</p>")

        with gr.Row():
            upscaler_name = gr.Radio(
                label='Upscaler Model',
                elem_id=f"{elem_id_prefix}_upscaler_name",
                choices=[x.name for x in shared.sd_upscalers],
                value=shared.sd_upscalers[0].name,
            )

        with gr.Row():
            tile_size = gr.Slider(
                elem_id=f"{elem_id_prefix}_tile_size",
                minimum=512,
                maximum=2048,
                step=64,
                label='Tile size',
                value=1024
            )
            feather_amount = gr.Slider(
                elem_id=f"{elem_id_prefix}_feather_amount",
                label='Feather Amount',
                minimum=0,
                maximum=100,
                step=1,
                value=50
            )
            overlap_percentage = gr.Slider(
                label='Overlap Percentage',
                minimum=0,
                maximum=50,
                step=1,
                value=20
            )
        with gr.Row():
            auto_adjust_tile_size = gr.Checkbox(
                label='Auto Adjust Tile Size',
                value=True
            )

        gr.HTML("<p>Prompt Generation Settings:</p>")
        with gr.Row():
            base_prompt = gr.Textbox(
                label="Prefix (Base Prompt)",
                placeholder="Enter a base prompt...",
                lines=1
            )
            clip_prompt_suffix = gr.Textbox(
                label="Suffix",
                placeholder="Enter a suffix for prompts...",
                lines=1
            )
        with gr.Row():
            prompt_methods = ["BLIP", "NONE"]
            prompt_method = gr.Dropdown(
                label="Prompt Generation Method",
                choices=prompt_methods,
                value="BLIP"
            )
        with gr.Row():
            prompt_mode_tile = gr.Dropdown(
                label="Prompt Complexity for Tiles",
                choices=["Simple", "Complex"],
                value="Simple"
            )

        with gr.Row():
            words_to_remove = gr.Textbox(
                label="Words to Remove from Tile Prompts",
                placeholder="Enter words to remove, separated by commas...",
                lines=1
            )
            categories_to_remove = gr.Textbox(
                label="Categories to Remove from Tile Prompts",
                placeholder="Enter categories to remove, separated by commas...",
                lines=1
            )

        gr.HTML("<p>Save options:</p>")
        with gr.Row():
            save_prompts_to_file = gr.Checkbox(
                label='Save Prompts to Excel File',
                value=True
            )
            upload_prompt_file = gr.File(
                label='Upload Prompt Excel File',
                file_types=['.xlsx']
            )

        with gr.Row():
            save_upscaled_image = gr.Checkbox(
                label="Save Upscaled Image",
                elem_id=f"{elem_id_prefix}_save_upscaled_image",
                value=True
            )
            save_tiles = gr.Checkbox(
                label="Save Each Tile",
                value=False
            )

        exclude_humans_animals = gr.Checkbox(
            label='Exclude Humans/Animals',
            value=False
        )

        return [
            upscaler_name,
            tile_size,
            feather_amount,
            overlap_percentage,
            auto_adjust_tile_size,
            base_prompt,
            clip_prompt_suffix,
            prompt_method,
            prompt_mode_tile,
            save_prompts_to_file,
            upload_prompt_file,
            save_upscaled_image,
            save_tiles,
            exclude_humans_animals,
            words_to_remove,
            categories_to_remove
        ]

    def run(self, p, *args):
        (
            upscaler_name,
            tile_size,
            feather_amount,
            overlap_percentage,
            auto_adjust_tile_size,
            base_prompt,
            clip_prompt_suffix,
            prompt_method,
            prompt_mode_tile,
            save_prompts_to_file,
            upload_prompt_file,
            save_upscaled_image,
            save_tiles,
            exclude_humans_animals,
            words_to_remove,
            categories_to_remove
        ) = args

        # Get upscaler index based on name
        upscaler_index = next((i for i, x in enumerate(shared.sd_upscalers) if x.name == upscaler_name), 0)

        # Assign the fixed seed immediately before any processing
        fix_seed(p)
        global_seed = p.seed

        # Initialize image
        init_img = p.init_images[0]
        if init_img is None:
            return Processed(p, [], p.seed, "Empty image")
        init_img = images.flatten(init_img, opts.img2img_background_color)

        # Upscaling process
        upscaler = USDUpscaler(
            p=p,
            image=init_img,
            upscaler_index=upscaler_index,
            save_redraw=save_upscaled_image,
            tile_size=tile_size,
            base_prompt=base_prompt,
            overlap_percentage=overlap_percentage,
            clip_prompt_suffix=clip_prompt_suffix,
            prompt_method=prompt_method,
            prompt_mode_tile=prompt_mode_tile,
            save_prompts_to_file=save_prompts_to_file,
            uploaded_prompt_file=upload_prompt_file,
            global_seed=global_seed,
            feather_amount=feather_amount,
            save_tiles=save_tiles,
            exclude_humans_animals=exclude_humans_animals,
            words_to_remove=words_to_remove,
            categories_to_remove=categories_to_remove,
            auto_adjust_tile_size=auto_adjust_tile_size
        )

        # Add extra info to display or save
        upscaler.add_extra_info()

        # Begin the upscaling and prompt processing
        upscaler.process()

        # Collect the result images
        result_images = upscaler.result_images

        return Processed(p, result_images, p.seed, upscaler.initial_info or "")
